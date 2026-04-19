# -*- coding: utf-8 -*-
"""routes/materials.py - 素材管理相关 API"""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse, Response

from db import get_db, now

router = APIRouter(tags=["materials"])

# ─── 常量 ───────────────────────────────────────────────────────────────
CATEGORY_ZH = {
    "quote":     "金句弹药库",
    "case":      "故事案例库",
    "viewpoint": "认知观点库",
    "action":    "实操行动库",
    "topic":     "IP选题映射",
}


# ─── 辅助函数 ───────────────────────────────────────────────────────────
# 移除代理字符范围 (U+D800 到 U+DFFF)
_SURROGATE_RANGE = {i: None for i in range(0xD800, 0xE000)}


def _clean_surrogate(obj):
    """递归清理字符串中的 Python 代理字符（JSON 不支持）"""
    if isinstance(obj, str):
        return obj.translate(_SURROGATE_RANGE)
    elif isinstance(obj, dict):
        return {k: _clean_surrogate(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_clean_surrogate(item) for item in obj]
    return obj


def _fetch_all_materials(source_id: str = "", category: str = "", starred_only: bool = False):
    """取全量素材，不分页"""
    conn = get_db()
    sql = """SELECT m.id, m.category, m.content, m.is_starred,
                    m.use_count, m.created_at, s.title as source_title
             FROM materials m LEFT JOIN sources s ON m.source_id = s.id
             WHERE 1=1"""
    params = []
    if source_id:
        sql += " AND m.source_id = %s"
        params.append(source_id)
    if category:
        sql += " AND m.category = %s"
        params.append(category)
    if starred_only:
        sql += " AND m.is_starred = 1"
    sql += " ORDER BY m.category, s.title, m.created_at"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ─── Routes ─────────────────────────────────────────────────────────────

# 获取素材列表
@router.get("/api/materials")
def list_materials(source_id: str = "", category: str = "", q: str = "",
                   starred: int = -1, review_only: int = 0, skip: int = 0, limit: int = 50):
    conn = get_db()
    sql = """SELECT m.*, s.title as source_title FROM materials m
             LEFT JOIN sources s ON m.source_id = s.id WHERE 1=1"""
    params = []
    if source_id:
        sql += " AND m.source_id=%s"
        params.append(source_id)
    if category:
        sql += " AND m.category=%s"
        params.append(category)
    if q:
        sql += " AND m.content LIKE %s"
        params.append(f"%{q}%")
    if starred == 1:
        sql += " AND m.is_starred=1"
    if review_only == 1:
        sql += " AND m.metadata LIKE '%_review_needed%'"
    sql += " ORDER BY m.created_at DESC LIMIT %s OFFSET %s"
    params += [limit, skip]
    rows = conn.execute(sql, params).fetchall()
    # 总数（去掉 ORDER BY 和 LIMIT/OFFSET）
    cnt_sql = sql.replace("SELECT m.*, s.title as source_title", "SELECT COUNT(*)")
    for keyword in ["ORDER BY", "LIMIT", "OFFSET"]:
        pos = cnt_sql.upper().rfind(keyword)
        if pos != -1:
            cnt_sql = cnt_sql[:pos].strip()
    total = conn.execute(cnt_sql, params[:-2]).fetchone()['count']
    conn.close()
    # 解析质量评分
    items = []
    for r in rows:
        d = dict(r)
        try:
            meta = json.loads(d.get("metadata") or "{}")
            d["_quality_score"] = meta.pop("_quality_score", None)
            d["_routing"] = meta.pop("_routing", None)
            d["_review_needed"] = meta.pop("_review_needed", False)
        except Exception:
            d["_quality_score"] = None
            d["_routing"] = None
            d["_review_needed"] = False
        items.append(d)
    return {"total": total, "items": _clean_surrogate(items)}


# 收藏/取消收藏
@router.post("/api/materials/{mid}/star")
def star_material(mid: str):
    conn = get_db()
    current = conn.execute("SELECT is_starred FROM materials WHERE id=%s", (mid,)).fetchone()
    if not current:
        raise HTTPException(404)
    new_val = 0 if current["is_starred"] else 1
    conn.execute("UPDATE materials SET is_starred=%s WHERE id=%s", (new_val, mid))
    conn.commit()
    conn.close()
    return {"is_starred": new_val}


# 记录使用次数
@router.post("/api/materials/{mid}/use")
def use_material(mid: str):
    conn = get_db()
    conn.execute("UPDATE materials SET use_count=use_count+1 WHERE id=%s", (mid,))
    conn.commit()
    conn.close()
    return {"ok": True}


# 删除素材
@router.delete("/api/materials/{mid}")
def delete_material(mid: str):
    conn = get_db()
    conn.execute("DELETE FROM materials WHERE id=%s", (mid,))
    conn.commit()
    conn.close()
    return {"ok": True}


# 获取素材质量评分详情
@router.get("/api/materials/{material_id}/quality")
def get_material_quality(material_id: str):
    """获取素材的质量评分详情"""
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM materials WHERE id=%s", (material_id,)
    ).fetchone()
    conn.close()

    if not row:
        raise HTTPException(404, "素材不存在")

    metadata = json.loads(row["metadata"] or "{}")
    quality_score = metadata.pop("_quality_score", {})

    return {
        "material_id": material_id,
        "quality_score": quality_score,
        "metadata": metadata,
        "suggestions": quality_score.get("suggestions", [])
    }


# ─── 导出 ───────────────────────────────────────────────────────────────

@router.get("/api/export/excel")
def export_excel(source_id: str = "", category: str = "", starred_only: bool = False):
    """导出素材为 Excel，每个分类一个 Sheet"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io
    from urllib.parse import quote

    rows = _fetch_all_materials(source_id, category, starred_only)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认空sheet

    # 按分类分组
    groups = defaultdict(list)
    for r in rows:
        groups[r["category"]].append(r)

    # 分类顺序
    order = ["quote", "case", "viewpoint", "action", "topic"]
    cats = order + [c for c in groups if c not in order]

    header_fill   = PatternFill("solid", fgColor="7C5CFC")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    star_fill     = PatternFill("solid", fgColor="FFF9E6")
    wrap_align    = Alignment(wrap_text=True, vertical="top")

    for cat in cats:
        items = groups.get(cat)
        if not items:
            continue
        sheet_name = CATEGORY_ZH.get(cat, cat)[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 表头
        headers = ["序号", "内容", "来源书籍", "收藏", "使用次数", "创建时间"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 22

        for ri, item in enumerate(items, 2):
            ws.cell(row=ri, column=1, value=ri - 1)
            ws.cell(row=ri, column=2, value=item["content"]).alignment = wrap_align
            ws.cell(row=ri, column=3, value=item.get("source_title", ""))
            ws.cell(row=ri, column=4, value="⭐" if item["is_starred"] else "")
            ws.cell(row=ri, column=5, value=item["use_count"])
            ws.cell(row=ri, column=6, value=(item["created_at"] or "")[:10])
            if item["is_starred"]:
                for ci in range(1, 7):
                    ws.cell(row=ri, column=ci).fill = star_fill

        # 列宽
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 6
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 12
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"IP军火库素材_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    encoded = quote(fname)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )


@router.get("/api/export/markdown")
def export_markdown(source_id: str = "", category: str = "", starred_only: bool = False):
    """导出素材为 Markdown 文本"""
    from urllib.parse import quote

    rows = _fetch_all_materials(source_id, category, starred_only)

    groups = defaultdict(list)
    for r in rows:
        groups[r["category"]].append(r)

    order = ["quote", "case", "viewpoint", "action", "topic"]
    cats = order + [c for c in groups if c not in order]

    lines = [
        f"# IP 军火库素材导出",
        f"> 导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        f"> 共 {len(rows)} 条素材\n",
    ]

    for cat in cats:
        items = groups.get(cat)
        if not items:
            continue
        lines.append(f"\n## {CATEGORY_ZH.get(cat, cat)}\n")
        # 按来源分组
        src_groups = defaultdict(list)
        for item in items:
            src_groups[item.get("source_title") or "未知来源"].append(item)
        for src_title, src_items in sorted(src_groups.items()):
            lines.append(f"\n### {src_title}\n")
            for idx, item in enumerate(src_items, 1):
                star = " ⭐" if item["is_starred"] else ""
                lines.append(f"{idx}. {item['content']}{star}\n")

    md_text = "\n".join(lines)
    fname = f"IP军火库素材_{datetime.now().strftime('%Y%m%d_%H%M')}.md"
    encoded = quote(fname)
    return Response(
        content=md_text.encode("utf-8"),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )
