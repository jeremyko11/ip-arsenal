# ─── UTF-8 编码 ──────────────────────────────────────────────────────────
import sys, os
os.environ.setdefault('PYTHONUTF8', '1')
os.environ.setdefault('PYTHONIOENCODING', 'utf-8')

# ─── 核心模块导入 ─────────────────────────────────────────────────────
from config import (
    BASE_DIR, DB_PATH, DATA_DIR, UPLOAD_DIR, FRONT_DIR,
    client, fallback_client, fallback2_client, get_minimax2_client,
    is_xunfei_blocked, ODL_AVAILABLE, get_paddle_ocr,
    MAX_CHARS, MAX_PROMPT_TEXT_CHARS, IP_DIRECTION,
    SMART_EXTRACTION_AVAILABLE,
)
from db import get_db, now
from ai import ai_extract
from tasks import enqueue_task, _start_workers, _recover_stuck_tasks
from services.process import process_source_task, process_source_task_smart
from extraction import (
    extract_text_from_epub, extract_text_from_txt, extract_text_from_docx,
    extract_text_from_pdf, extract_text_from_url,
)
from prompts import EXTRACT_MODES, build_prompt, parse_materials, parse_atomic_notes

# ─── 路由模块 ─────────────────────────────────────────────────────────
from routes import sources, materials, wechat, pushutree, media, misc

# ─── 标准库（供路由和 process_source_task 使用）──────────────────────
import json, sqlite3, uuid, time, asyncio, base64, io, threading
import queue
from datetime import datetime
from pathlib import Path
from typing import Optional, List

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

# ─── FastAPI 应用 ────────────────────────────────────────────────────
from fastapi.responses import ORJSONResponse
try:
    import orjson
    class _UnicodeJSONResponse(ORJSONResponse):
        pass
    _json_response_class = _UnicodeJSONResponse
except ImportError:
    from fastapi.responses import Response as _Resp
    class _UnicodeJSONResponse(_Resp):
        media_type = 'application/json; charset=utf-8'
        def render(self, content) -> bytes:
            return json.dumps(content, ensure_ascii=False, allow_nan=False).encode('utf-8')
    _json_response_class = _UnicodeJSONResponse

app = FastAPI(title='IP Arsenal API', version='1.0.0', default_response_class=_json_response_class)
app.add_middleware(CORSMiddleware, allow_origins=['*'], allow_methods=['*'], allow_headers=['*'])

# ─── 路由注册 ─────────────────────────────────────────────────────────
app.include_router(sources.router)
app.include_router(materials.router)
app.include_router(wechat.router)
app.include_router(pushutree.router)
app.include_router(media.router)
app.include_router(misc.router)

@app.on_event('startup')
def on_startup():
    _start_workers()
    _recover_stuck_tasks()

# ─── Worker 队列引用（已迁移到 tasks.py）────────────────────────────
from tasks import _task_queue, WORKER_COUNT, TASK_MAX_SECONDS, _worker_threads, _worker_heartbeats
from tasks import _worker_loop, _spawn_worker, _watchdog_loop

# ─── 以下为占位说明 ──────────────────────────────────────────────────
# config.py    → 常量、AI客户端初始化
# db.py        → 数据库初始化 (init_db, migrate_db, get_db, now)
# ai.py        → AI调用链 (ai_extract)
# tasks.py     → Worker线程池 (enqueue_task, _start_workers, _recover_stuck_tasks)
# extraction.py → PDF/EPUB/TXT/DOCX/URL 文本提取
# prompts.py   → 提示词构建 (build_prompt, parse_materials, EXTRACT_MODES)


def parse_materials(source_id: str, raw_content: str) -> List[dict]:
    """将AI输出解析为结构化素材条目"""
    materials = []
    now_str = now()

    lines = raw_content.split("\n")
    current_cat = "quote"
    current_block = []

    cat_map = {
        "第一部分": "quote",
        "金句弹药库": "quote",
        "第二部分": "case",
        "故事与案例": "case",
        "第三部分": "viewpoint",
        "认知与观点": "viewpoint",
        "第四部分": "action",
        "实操与行动": "action",
        "第五部分": "topic",
        "IP选题映射": "topic",
        "书籍综合评级": "rating",
    }

    def flush_block():
        nonlocal current_block
        text = "\n".join(current_block).strip()
        if text and len(text) > 10:
            meta = {}
            # 提取风险标签
            for line in current_block:
                if "风险标签" in line:
                    if "✅安全" in line: meta["risk"] = "safe"
                    elif "⚠️需语境" in line: meta["risk"] = "context"
                    elif "❌禁用" in line: meta["risk"] = "forbidden"
                if "爆点场景" in line:
                    if "🔥爆款" in line: meta["scene"] = "viral"
                    elif "💡治愈" in line: meta["scene"] = "heal"
                    elif "📚深度" in line: meta["scene"] = "deep"
                if "改写成本" in line:
                    if "⚡0成本" in line: meta["cost"] = "zero"
                    elif "✂️中成本" in line: meta["cost"] = "mid"
                    elif "🛠️高成本" in line: meta["cost"] = "high"
                if "时效熔断" in line:
                    if "✅长效" in line: meta["timeliness"] = "long"
                    elif "⚠️需更新" in line: meta["timeliness"] = "update"
                    elif "❌已过期" in line: meta["timeliness"] = "expired"

            materials.append({
                "id": str(uuid.uuid4()),
                "source_id": source_id,
                "category": current_cat,
                "content": text,
                "metadata": json.dumps(meta, ensure_ascii=False),
                "tags": "[]",
                "platform": "[]",
                "use_count": 0,
                "is_starred": 0,
                "created_at": now_str,
            })
        current_block = []

    for line in lines:
        # 检测分类标题
        new_cat = None
        for key, cat in cat_map.items():
            if key in line and ("##" in line or "【" in line):
                new_cat = cat
                break

        if new_cat:
            flush_block()
            current_cat = new_cat
        else:
            # 金句单条切割（以 > 开头的引用块）
            if current_cat == "quote" and line.startswith("> ") and "风险标签" not in line and "爆点场景" not in line and "改写成本" not in line and "时效熔断" not in line:
                if current_block:
                    flush_block()
                current_block = [line]
            elif current_cat == "quote" and line.startswith("> "):
                current_block.append(line)
            # 案例/观点/行动 — 以 ** 开头新条目
            elif current_cat in ("case","viewpoint","action","topic") and line.startswith("**") and line.endswith("**"):
                flush_block()
                current_block = [line]
            elif current_cat in ("case","viewpoint","action","topic") and line.startswith("- ") and current_block:
                current_block.append(line)
            elif current_cat == "topic" and line.strip().startswith(("1.","2.","3.","4.","5.","6.","7.","8.","9.","10.","11.","12.","13.","14.","15.")):
                flush_block()
                current_block = [line]
            elif current_block:
                current_block.append(line)

    flush_block()
    return materials


# ─── process_source_task 和 process_source_task_smart 已迁移到 services/process.py ─────
# ─── API 路由已拆分到 routes/ 目录（sources.py / materials.py / wechat.py / pushutree.py / media.py / misc.py）─────



    file_id = str(uuid.uuid4())
    filename = file.filename or "unknown.pdf"
    save_path = UPLOAD_DIR / f"{file_id}_{filename}"

    content = await file.read()
    save_path.write_bytes(content)

    source_id = str(uuid.uuid4())
    task_id = str(uuid.uuid4())
    title = Path(filename).stem

    # 根据文件后缀决定 source type
    ext = Path(filename).suffix.lower()
    if ext == ".epub":
        source_type = "epub"
    elif ext in (".txt", ".md"):
        source_type = "txt"
    elif ext == ".docx":
        source_type = "docx"
    else:
        source_type = "book"  # pdf / 其他

    conn = get_db()
    conn.execute(
        "INSERT INTO sources VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (source_id, title, source_type, str(save_path), None, "[]", 0, 0, "pending", None, 0, now(), now())
    )
    conn.execute(
        "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?)",
        (task_id, source_id, "pending", 0, "等待处理...", None, now(), now())
    )
    conn.commit()
    conn.close()

    enqueue_task(task_id, source_id, mode)
    return {"source_id": source_id, "task_id": task_id, "title": title}


# ─── 文件夹批量导入 ───────────────────────────────────────────────────

class ImportFolderRequest(BaseModel):
    folder_path: str
    mode: str = "full"
    recursive: bool = False        # 是否递归子目录
    skip_existing: bool = True     # 跳过已导入的同名书籍


@app.post("/api/sources/import-folder")
async def import_folder(req: ImportFolderRequest):
    """扫描本地文件夹，将所有 PDF/EPUB/TXT 加入书库并启动提炼任务"""
    folder = Path(req.folder_path.strip())
    if not folder.exists():
        raise HTTPException(400, f"路径不存在：{folder}")
    if not folder.is_dir():
        raise HTTPException(400, f"该路径不是文件夹：{folder}")

    # 扫描 PDF + EPUB + TXT + DOCX
    SUPPORTED_EXTS = (".pdf", ".epub", ".txt", ".md", ".docx")
    if req.recursive:
        all_files = [f for f in sorted(folder.rglob("*")) if f.suffix.lower() in SUPPORTED_EXTS]
    else:
        all_files = [f for f in sorted(folder.glob("*")) if f.suffix.lower() in SUPPORTED_EXTS]

    if not all_files:
        return {"total": 0, "queued": 0, "skipped": 0, "tasks": [],
                "message": "文件夹中没有找到 PDF/EPUB/TXT 文件"}

    # 查已有书名，跳过重复
    conn = get_db()
    existing_titles = set()
    if req.skip_existing:
        rows = conn.execute("SELECT title FROM sources WHERE type IN ('book','epub','txt')").fetchall()
        existing_titles = {r["title"] for r in rows}

    tasks_created = []
    skipped = []

    for file_path in all_files:
        title = file_path.stem
        if req.skip_existing and title in existing_titles:
            skipped.append(title)
            continue

        # 根据后缀设置 type
        ext = file_path.suffix.lower()
        if ext == ".epub":
            source_type = "epub"
        elif ext in (".txt", ".md"):
            source_type = "txt"
        elif ext == ".docx":
            source_type = "docx"
        else:
            source_type = "book"

        # 复制到 uploads
        file_id  = str(uuid.uuid4())
        dest     = UPLOAD_DIR / f"{file_id}_{file_path.name}"
        try:
            import shutil
            shutil.copy2(str(file_path), str(dest))
        except Exception as e:
            skipped.append(f"{title} (复制失败: {e})")
            continue

        source_id = str(uuid.uuid4())
        task_id   = str(uuid.uuid4())
        conn.execute(
            "INSERT INTO sources VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (source_id, title, source_type, str(dest), None, "[]", 0, 0,
             "pending", None, 0, now(), now())
        )
        conn.execute(
            "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?)",
            (task_id, source_id, "pending", 0, "等待批量提炼...", None, now(), now())
        )
        tasks_created.append({"task_id": task_id, "source_id": source_id, "title": title})
        existing_titles.add(title)   # 同一批次也去重

    conn.commit()
    conn.close()

    # 入队（由全局 Worker 线程池串行消费，避免同时打爆 API）
    for t in tasks_created:
        enqueue_task(t["task_id"], t["source_id"], req.mode)

    return {
        "total":   len(all_files),
        "queued":  len(tasks_created),
        "skipped": len(skipped),
        "tasks":   tasks_created,
        "message": f"已加入队列 {len(tasks_created)} 本，跳过 {len(skipped)} 本"
    }


@app.get("/api/sources/scan-folder")
def scan_folder(path: str, recursive: bool = False):
    """预扫描文件夹，返回文件列表（不导入），用于前端预览"""
    folder = Path(path.strip())
    if not folder.exists() or not folder.is_dir():
        raise HTTPException(400, "路径无效或不是文件夹")

    # 扫描所有支持的文件类型
    SUPPORTED_EXTS = (".pdf", ".epub", ".txt", ".md", ".docx")
    if recursive:
        all_files = sorted(f for f in folder.rglob("*") if f.suffix.lower() in SUPPORTED_EXTS)
    else:
        all_files = sorted(f for f in folder.glob("*") if f.suffix.lower() in SUPPORTED_EXTS)

    conn = get_db()
    existing_titles = {r["title"] for r in
                       conn.execute("SELECT title FROM sources WHERE type IN ('book','epub','txt','docx')").fetchall()}
    conn.close()

    items = []
    for p in all_files[:200]:   # 预览最多200条
        size_mb = round(p.stat().st_size / 1024 / 1024, 1)
        items.append({
            "name": p.stem,
            "filename": p.name,
            "size_mb": size_mb,
            "already_imported": p.stem in existing_titles
        })

    return {"path": str(folder), "count": len(all_files), "items": items}


@app.get("/api/pick-folder")
def pick_folder(initial_dir: str = ""):
    """弹出系统原生文件夹选择对话框，返回用户选中的路径。
    使用 tkinter（Python 标准库），无需额外安装。
    在后台线程中运行以避免阻塞 asyncio 事件循环。
    """
    import threading

    result_holder = {"path": None, "error": None}

    def _show_dialog():
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()          # 隐藏主窗口
            root.attributes("-topmost", True)   # 置顶对话框
            root.update()
            kwargs = {"title": "选择书籍文件夹"}
            if initial_dir and Path(initial_dir).is_dir():
                kwargs["initialdir"] = initial_dir
            selected = filedialog.askdirectory(**kwargs)
            root.destroy()
            result_holder["path"] = selected or ""
        except Exception as e:
            result_holder["error"] = str(e)

    t = threading.Thread(target=_show_dialog, daemon=True)
    t.start()
    t.join(timeout=120)   # 最多等 2 分钟

    if result_holder["error"]:
        raise HTTPException(500, f"无法打开文件夹选择框：{result_holder['error']}")

    return {"selected": result_holder["path"] or ""}


# 添加文字/URL来源
class AddSourceRequest(BaseModel):
    title: str
    type: str   # text / url
    content: str
    mode: str = "full"

@app.post("/api/sources/add")
async def add_source(req: AddSourceRequest):
    source_id = str(uuid.uuid4())
    task_id = str(uuid.uuid4())

    conn = get_db()
    conn.execute(
        "INSERT INTO sources VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (source_id, req.title, req.type, None, req.content, "[]", 0, 0, "pending", None, 0, now(), now())
    )
    conn.execute(
        "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?)",
        (task_id, source_id, "pending", 0, "等待处理...", None, now(), now())
    )
    conn.commit()
    conn.close()

    enqueue_task(task_id, source_id, req.mode)
    return {"source_id": source_id, "task_id": task_id}

# 查询任务进度
@app.get("/api/tasks/{task_id}")
def get_task(task_id: str):
    conn = get_db()
    task = conn.execute("SELECT * FROM tasks WHERE id=?", (task_id,)).fetchone()
    conn.close()
    if not task:
        raise HTTPException(404, "任务不存在")
    return dict(task)

# 获取所有书库
@app.get("/api/sources")
def list_sources(q: str = "", type: str = ""):
    conn = get_db()
    sql = "SELECT s.*, (SELECT COUNT(*) FROM materials m WHERE m.source_id=s.id) as material_count FROM sources s WHERE 1=1"
    params = []
    if q:
        sql += " AND s.title LIKE ?"
        params.append(f"%{q}%")
    if type:
        sql += " AND s.type=?"
        params.append(type)
    sql += " ORDER BY s.created_at DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# 重试失败任务
@app.post("/api/sources/{source_id}/retry")
async def retry_source(source_id: str, mode: str = "full"):
    """重新提炼：重置 source + task 状态，立即重新入队"""
    conn = get_db()
    src = conn.execute("SELECT * FROM sources WHERE id=?", (source_id,)).fetchone()
    if not src:
        conn.close()
        raise HTTPException(404, "书籍不存在")

    # 重置 source 状态
    conn.execute(
        "UPDATE sources SET status='pending', error_msg=NULL, updated_at=? WHERE id=?",
        (now(), source_id)
    )

    # 复用已有 task（最新的那条），或新建一条
    existing_task = conn.execute(
        "SELECT id FROM tasks WHERE source_id=? ORDER BY created_at DESC LIMIT 1",
        (source_id,)
    ).fetchone()
    if existing_task:
        task_id = existing_task["id"]
        conn.execute(
            "UPDATE tasks SET status='pending', progress=0, message='重试中，等待处理...', result=NULL, updated_at=? WHERE id=?",
            (now(), task_id)
        )
    else:
        task_id = str(uuid.uuid4())
        conn.execute(
            "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?)",
            (task_id, source_id, "pending", 0, "重试中，等待处理...", None, now(), now())
        )

    conn.commit()
    conn.close()

    # 立即重新入队
    enqueue_task(task_id, source_id, mode)
    return {"ok": True, "task_id": task_id, "status": "pending"}


# 一键恢复所有卡住的任务
@app.post("/api/sources/recover-all")
def recover_all_stuck():
    """将所有 processing/pending 状态的任务重新入队"""
    conn = get_db()
    try:
        # 先将 processing 重置为 pending
        conn.execute(
            "UPDATE sources SET status='pending', updated_at=? WHERE status='processing'",
            (now(),)
        )
        conn.execute(
            "UPDATE tasks SET status='pending', message='等待恢复处理...', updated_at=? WHERE status='processing'",
            (now(),)
        )
        conn.commit()

        rows = conn.execute(
            """SELECT t.id as task_id, t.source_id
               FROM tasks t JOIN sources s ON t.source_id = s.id
               WHERE s.status = 'pending'
               ORDER BY s.created_at"""
        ).fetchall()

        queued = 0
        for row in rows:
            enqueue_task(row["task_id"], row["source_id"], "full")
            queued += 1

        return {"ok": True, "queued": queued, "queue_size": _task_queue.qsize()}
    finally:
        conn.close()



# ─── AI Studio OCR Token 配置接口 ───────────────────────────────────
class AistudioTokenRequest(BaseModel):
    token: str

@app.get("/api/ocr/aistudio-token")
def get_aistudio_token():
    """查询当前 AI Studio token 状态（不返回 token 明文）"""
    return {"configured": bool(AISTUDIO_TOKEN), "length": len(AISTUDIO_TOKEN)}

@app.post("/api/ocr/aistudio-token")
def set_aistudio_token(req: AistudioTokenRequest):
    """设置 AI Studio OCR token（运行时生效，重启后失效，如需持久化请设置环境变量 AISTUDIO_TOKEN）"""
    global AISTUDIO_TOKEN
    AISTUDIO_TOKEN = req.token.strip()
    os.environ["AISTUDIO_TOKEN"] = AISTUDIO_TOKEN
    return {"ok": True, "configured": bool(AISTUDIO_TOKEN)}


# ─── AI 模型选择接口 ─────────────────────────────────────────────────
class AiModelRequest(BaseModel):
    preferred: str          # "xunfei" | "minimax" | "minimax2" | "deepseek"
    minimax2_key: Optional[str] = None   # 若切换 minimax2，可同时传入 key
    minimax2_base: Optional[str] = None  # MiniMax2 API Base（可选，默认官方直连）

@app.get("/api/ai-model")
def get_ai_model():
    """查询当前 AI 首选模型状态"""
    global _AI_PREFERRED, MINIMAX_API_KEY_2, MINIMAX_API_BASE_2
    labels = {
        "xunfei":   "讯飞星辰",
        "minimax":  "MiniMax（账号1）",
        "minimax2": "MiniMax（账号2）",
        "deepseek": "DeepSeek",
    }
    return {
        "preferred": _AI_PREFERRED,
        "preferred_label": labels.get(_AI_PREFERRED, _AI_PREFERRED),
        "minimax2_configured": bool(MINIMAX_API_KEY_2),
        "minimax2_key_preview": (MINIMAX_API_KEY_2[:8] + "****") if MINIMAX_API_KEY_2 else "",
        "minimax2_base": MINIMAX_API_BASE_2,
        "available_models": [
            {"id": "xunfei",   "label": "讯飞星辰",       "note": "每日 5000万 token，文本理解最强"},
            {"id": "minimax",  "label": "MiniMax（账号1）","note": "每6小时 4000次，旧账号"},
            {"id": "minimax2", "label": "MiniMax（账号2）","note": "每6小时 4000次，需填入 API Key"},
            {"id": "deepseek", "label": "DeepSeek",        "note": "最终兜底，按量计费"},
        ]
    }

@app.post("/api/ai-model")
def set_ai_model(req: AiModelRequest):
    """切换全局 AI 首选模型（立即生效，后续所有提炼任务从该模型开始）"""
    global _AI_PREFERRED, MINIMAX_API_KEY_2, MINIMAX_API_BASE_2, _minimax2_client
    allowed = {"xunfei", "minimax", "minimax2", "deepseek"}
    if req.preferred not in allowed:
        raise HTTPException(status_code=400, detail=f"preferred 必须是: {allowed}")
    
    # 如果切换到 minimax2，且传入了 key，则更新
    if req.minimax2_key is not None:
        MINIMAX_API_KEY_2 = req.minimax2_key.strip()
        _minimax2_client = None   # 清除缓存，下次重新创建
        if req.minimax2_base:
            MINIMAX_API_BASE_2 = req.minimax2_base.strip()
    
    # 切 minimax2 但 key 还没有
    if req.preferred == "minimax2" and not MINIMAX_API_KEY_2:
        raise HTTPException(status_code=400, detail="切换到 MiniMax 账号2 前，请先在 minimax2_key 字段填写 API Key")
    
    _AI_PREFERRED = req.preferred
    labels = {"xunfei":"讯飞星辰","minimax":"MiniMax（账号1）","minimax2":"MiniMax（账号2）","deepseek":"DeepSeek"}
    print(f"[AI] 用户切换首选模型 → {labels.get(_AI_PREFERRED)}")
    return {"ok": True, "preferred": _AI_PREFERRED, "label": labels.get(_AI_PREFERRED)}


# 查询正在处理中的任务详情（供前端实时进度面板使用）
@app.get("/api/processing-tasks")
def processing_tasks():
    """返回所有 processing 状态的书籍及其最新任务进度（title、progress、message）"""
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT s.id as source_id, s.title, s.status as source_status,
                   t.id as task_id, t.status as task_status,
                   t.progress, t.message, t.updated_at
            FROM sources s
            LEFT JOIN tasks t ON t.source_id = s.id
            WHERE s.status IN ('processing', 'pending')
            ORDER BY t.updated_at DESC
        """).fetchall()
        # 每个 source 只取最新一条 task
        seen = set()
        result = []
        for row in rows:
            sid = row["source_id"]
            if sid in seen:
                continue
            seen.add(sid)
            result.append({
                "source_id": sid,
                "title": row["title"],
                "source_status": row["source_status"],
                "task_id": row["task_id"],
                "task_status": row["task_status"],
                "progress": row["progress"] or 0,
                "message": row["message"] or "",
                "updated_at": row["updated_at"] or "",
            })
        return result
    finally:
        conn.close()


# 查询当前队列状态
@app.get("/api/queue-status")
def queue_status():
    """返回当前任务队列大小 + worker 线程存活状态 + 心跳信息"""
    conn = get_db()
    try:
        pending = conn.execute("SELECT COUNT(*) FROM sources WHERE status='pending'").fetchone()[0]
        processing = conn.execute("SELECT COUNT(*) FROM sources WHERE status='processing'").fetchone()[0]
        done = conn.execute("SELECT COUNT(*) FROM sources WHERE status='done'").fetchone()[0]
        error = conn.execute("SELECT COUNT(*) FROM sources WHERE status='error'").fetchone()[0]
        alive_workers = sum(1 for t in _worker_threads if t.is_alive())
        # 心跳信息：0=idle, >0=正在处理，值为开始时间戳
        now_ts = time.time()
        worker_states = []
        for i, t in enumerate(_worker_threads):
            wid = i + 1
            hb = _worker_heartbeats.get(wid, 0)
            if not t.is_alive():
                state = "dead"
            elif hb == 0:
                state = "idle"
            else:
                elapsed = int(now_ts - hb)
                state = f"busy({elapsed}s)"
            worker_states.append(state)
        return {
            "queue_size": _task_queue.qsize(),
            "workers": WORKER_COUNT,
            "alive_workers": alive_workers,
            "worker_states": worker_states,
            "pending": pending,
            "processing": processing,
            "done": done,
            "error": error,
        }
    finally:
        conn.close()


# 手动重启 workers（出现死线程时应急用）
@app.post("/api/workers/restart")
def restart_workers():
    """手动重建所有 worker 线程（用于 worker 线程异常死亡时应急恢复）"""
    global _worker_threads
    dead = [i for i, t in enumerate(_worker_threads) if not t.is_alive()]
    for i in dead:
        new_t = _spawn_worker(i + 1)
        _worker_threads[i] = new_t
        print(f"[API] Worker-{i+1} 手动重建")
    alive = sum(1 for t in _worker_threads if t.is_alive())
    return {"ok": True, "restarted": len(dead), "alive_workers": alive, "queue_size": _task_queue.qsize()}


# 强制跳过卡住的 processing 任务（将其标记为 error，重新入队 pending 任务）
@app.post("/api/sources/skip-stuck")
def skip_stuck_tasks():
    """将所有长时间卡在 processing 的任务标记为 error，让队列继续跑"""
    conn = get_db()
    try:
        # 找出 processing 超过 10 分钟的任务（超时阈值宽松，避免误杀正常任务）
        stuck_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        rows = conn.execute(
            """SELECT s.id as source_id, t.id as task_id 
               FROM sources s LEFT JOIN tasks t ON t.source_id = s.id
               WHERE s.status = 'processing'"""
        ).fetchall()
        skipped = 0
        for row in rows:
            conn.execute("UPDATE sources SET status='error',error_msg=?,updated_at=? WHERE id=?",
                         ("手动强制跳过（任务卡住）", now(), row["source_id"]))
            if row["task_id"]:
                conn.execute("UPDATE tasks SET status='error',message=?,updated_at=? WHERE id=?",
                             ("手动强制跳过（任务卡住）", now(), row["task_id"]))
            skipped += 1
        conn.commit()
        print(f"[API] 强制跳过 {skipped} 个卡住任务")
        return {"ok": True, "skipped": skipped, "queue_size": _task_queue.qsize()}
    finally:
        conn.close()


# 重置所有错误任务为 pending，重新入队
@app.post("/api/sources/retry-errors")
def retry_error_tasks():
    """将所有 error 状态的任务重置为 pending 重新入队"""
    conn = get_db()
    try:
        rows = conn.execute(
            """SELECT t.id as task_id, t.source_id
               FROM tasks t JOIN sources s ON t.source_id = s.id
               WHERE s.status = 'error'"""
        ).fetchall()
        retried = 0
        for row in rows:
            conn.execute("UPDATE sources SET status='pending',error_msg=NULL,updated_at=? WHERE id=?",
                         (now(), row["source_id"]))
            conn.execute("UPDATE tasks SET status='pending',message='重新等待处理...',progress=0,updated_at=? WHERE id=?",
                         (now(), row["task_id"]))
            enqueue_task(row["task_id"], row["source_id"], "full")
            retried += 1
        conn.commit()
        print(f"[API] 重试 {retried} 个错误任务")
        return {"ok": True, "retried": retried, "queue_size": _task_queue.qsize()}
    finally:
        conn.close()






# 一键清空书库（全部删除）
@app.delete("/api/sources")
def delete_all_sources():
    conn = get_db()
    try:
        # 先取出所有文件路径，用于删除物理文件
        rows = conn.execute("SELECT file_path FROM sources WHERE file_path IS NOT NULL AND file_path != ''").fetchall()
        for row in rows:
            try:
                Path(row["file_path"]).unlink(missing_ok=True)
            except Exception:
                pass
        conn.execute("DELETE FROM materials")
        conn.execute("DELETE FROM tasks")
        conn.execute("DELETE FROM sources")
        conn.commit()
        return {"ok": True, "deleted": len(rows)}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# 删除来源
@app.delete("/api/sources/{source_id}")
def delete_source(source_id: str):
    conn = get_db()
    conn.execute("DELETE FROM materials WHERE source_id=?", (source_id,))
    conn.execute("DELETE FROM tasks WHERE source_id=?", (source_id,))
    src = conn.execute("SELECT file_path FROM sources WHERE id=?", (source_id,)).fetchone()
    if src and src["file_path"]:
        try: Path(src["file_path"]).unlink()
        except: pass
    conn.execute("DELETE FROM sources WHERE id=?", (source_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

# 获取素材列表
@app.get("/api/materials")
def list_materials(source_id: str = "", category: str = "", q: str = "",
                   starred: int = -1, review_only: int = 0, skip: int = 0, limit: int = 50):
    conn = get_db()
    sql = """SELECT m.*, s.title as source_title FROM materials m
             LEFT JOIN sources s ON m.source_id = s.id WHERE 1=1"""
    params = []
    if source_id:
        sql += " AND m.source_id=?"
        params.append(source_id)
    if category:
        sql += " AND m.category=?"
        params.append(category)
    if q:
        sql += " AND m.content LIKE ?"
        params.append(f"%{q}%")
    if starred == 1:
        sql += " AND m.is_starred=1"
    if review_only == 1:
        sql += " AND m.metadata LIKE '%_review_needed%'"
    sql += " ORDER BY m.created_at DESC LIMIT ? OFFSET ?"
    params += [limit, skip]
    rows = conn.execute(sql, params).fetchall()
    # 总数
    cnt_sql = sql.replace("SELECT m.*, s.title as source_title", "SELECT COUNT(*)")
    cnt_sql = cnt_sql[:cnt_sql.rfind("LIMIT")]
    total = conn.execute(cnt_sql, params[:-2]).fetchone()[0]
    conn.close()
    # 解析质量评分
    items = []
    for r in rows:
        d = dict(r)
        try:
            meta = json.loads(d.get("metadata") or "{}")
            d["_quality_score"] = meta.pop("_quality_score", None)
            d["_routing"] = meta.pop("_routing", None)
            d["_review_needed"] = meta.pop("_review_needed", False)
        except Exception:
            d["_quality_score"] = None
            d["_routing"] = None
            d["_review_needed"] = False
        items.append(d)
    return {"total": total, "items": items}

# 收藏/取消收藏
@app.post("/api/materials/{mid}/star")
def star_material(mid: str):
    conn = get_db()
    current = conn.execute("SELECT is_starred FROM materials WHERE id=?", (mid,)).fetchone()
    if not current:
        raise HTTPException(404)
    new_val = 0 if current["is_starred"] else 1
    conn.execute("UPDATE materials SET is_starred=? WHERE id=?", (new_val, mid))
    conn.commit()
    conn.close()
    return {"is_starred": new_val}

# 记录使用次数
@app.post("/api/materials/{mid}/use")
def use_material(mid: str):
    conn = get_db()
    conn.execute("UPDATE materials SET use_count=use_count+1 WHERE id=?", (mid,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/materials/{mid}")
def delete_material(mid: str):
    conn = get_db()
    conn.execute("DELETE FROM materials WHERE id=?", (mid,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ─── 导出 ───────────────────────────────────────────────────────────

CATEGORY_ZH = {
    "quote":     "金句弹药库",
    "case":      "故事案例库",
    "viewpoint": "认知观点库",
    "action":    "实操行动库",
    "topic":     "IP选题映射",
}

def _fetch_all_materials(source_id: str = "", category: str = "", starred_only: bool = False):
    """取全量素材，不分页"""
    conn = get_db()
    sql = """SELECT m.id, m.category, m.content, m.is_starred,
                    m.use_count, m.created_at, s.title as source_title
             FROM materials m LEFT JOIN sources s ON m.source_id = s.id
             WHERE 1=1"""
    params = []
    if source_id:
        sql += " AND m.source_id = ?"
        params.append(source_id)
    if category:
        sql += " AND m.category = ?"
        params.append(category)
    if starred_only:
        sql += " AND m.is_starred = 1"
    sql += " ORDER BY m.category, s.title, m.created_at"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/export/excel")
def export_excel(source_id: str = "", category: str = "", starred_only: bool = False):
    """导出素材为 Excel，每个分类一个 Sheet"""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io

    rows = _fetch_all_materials(source_id, category, starred_only)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认空sheet

    # 按分类分组
    from collections import defaultdict
    groups = defaultdict(list)
    for r in rows:
        groups[r["category"]].append(r)

    # 分类顺序
    order = ["quote", "case", "viewpoint", "action", "topic"]
    cats = order + [c for c in groups if c not in order]

    header_fill   = PatternFill("solid", fgColor="7C5CFC")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    star_fill     = PatternFill("solid", fgColor="FFF9E6")
    wrap_align    = Alignment(wrap_text=True, vertical="top")

    for cat in cats:
        items = groups.get(cat)
        if not items:
            continue
        sheet_name = CATEGORY_ZH.get(cat, cat)[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 表头
        headers = ["序号", "内容", "来源书籍", "收藏", "使用次数", "创建时间"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 22

        for ri, item in enumerate(items, 2):
            ws.cell(row=ri, column=1, value=ri - 1)
            ws.cell(row=ri, column=2, value=item["content"]).alignment = wrap_align
            ws.cell(row=ri, column=3, value=item.get("source_title", ""))
            ws.cell(row=ri, column=4, value="⭐" if item["is_starred"] else "")
            ws.cell(row=ri, column=5, value=item["use_count"])
            ws.cell(row=ri, column=6, value=(item["created_at"] or "")[:10])
            if item["is_starred"]:
                for ci in range(1, 7):
                    ws.cell(row=ri, column=ci).fill = star_fill

        # 列宽
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 6
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 12
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    fname = f"IP军火库素材_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    encoded = quote(fname)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )


@app.get("/api/export/markdown")
def export_markdown(source_id: str = "", category: str = "", starred_only: bool = False):
    """导出素材为 Markdown 文本"""
    from fastapi.responses import Response
    from urllib.parse import quote

    rows = _fetch_all_materials(source_id, category, starred_only)

    from collections import defaultdict
    groups = defaultdict(list)
    for r in rows:
        groups[r["category"]].append(r)

    order = ["quote", "case", "viewpoint", "action", "topic"]
    cats = order + [c for c in groups if c not in order]

    lines = [
        f"# IP 军火库素材导出",
        f"> 导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        f"> 共 {len(rows)} 条素材\n",
    ]

    for cat in cats:
        items = groups.get(cat)
        if not items:
            continue
        lines.append(f"\n## {CATEGORY_ZH.get(cat, cat)}\n")
        # 按来源分组
        src_groups = defaultdict(list)
        for item in items:
            src_groups[item.get("source_title") or "未知来源"].append(item)
        for src_title, src_items in sorted(src_groups.items()):
            lines.append(f"\n### 📖 {src_title}\n")
            for idx, item in enumerate(src_items, 1):
                star = " ⭐" if item["is_starred"] else ""
                lines.append(f"{idx}. {item['content']}{star}\n")

    md_text = "\n".join(lines)
    fname = f"IP军火库素材_{datetime.now().strftime('%Y%m%d_%H%M')}.md"
    encoded = quote(fname)
    return Response(
        content=md_text.encode("utf-8"),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )


# ─── 公众号排版 ──────────────────────────────────────────────────────

# 主题名称→中文标签映射（用于返回给前端选择器展示）
WECHAT_THEMES = [
    # 深度长文
    {"id": "newspaper",    "label": "报纸风",    "group": "深度长文", "desc": "《纽约时报》蓝灰，适合新闻/分析"},
    {"id": "magazine",     "label": "杂志风",    "group": "深度长文", "desc": "高端杂志版式，适合品质内容"},
    {"id": "coffee-house", "label": "咖啡馆",    "group": "深度长文", "desc": "咖啡色暖调，适合随笔/日记"},
    {"id": "ink",          "label": "水墨",      "group": "深度长文", "desc": "极简黑白，适合文学/散文"},
    # 科技产品
    {"id": "bytedance",    "label": "字节风",    "group": "科技产品", "desc": "蓝绿渐变，现代科技感"},
    {"id": "github",       "label": "GitHub",    "group": "科技产品", "desc": "开发者风格，代码友好"},
    {"id": "sspai",        "label": "少数派",    "group": "科技产品", "desc": "少数派橙色，适合效率/工具"},
    {"id": "midnight",     "label": "深夜",      "group": "科技产品", "desc": "深色背景，赛博朋克感"},
    # 文艺随笔
    {"id": "terracotta",   "label": "陶土",      "group": "文艺随笔", "desc": "暖橙圆角，轻松活泼"},
    {"id": "mint-fresh",   "label": "薄荷绿",    "group": "文艺随笔", "desc": "清新绿色，适合生活/美食"},
    {"id": "sunset-amber", "label": "琥珀",      "group": "文艺随笔", "desc": "琥珀暖色，情感/故事"},
    {"id": "lavender-dream","label": "薰衣草",   "group": "文艺随笔", "desc": "紫色梦幻，适合情感内容"},
    # 活力动态
    {"id": "sports",       "label": "运动",      "group": "活力动态", "desc": "渐变条纹，活力满满"},
    {"id": "bauhaus",      "label": "包豪斯",    "group": "活力动态", "desc": "几何色块，大胆设计感"},
    {"id": "chinese",      "label": "朱红国风",  "group": "活力动态", "desc": "朱红古典，传统文化"},
    {"id": "wechat-native","label": "微信原生",  "group": "活力动态", "desc": "微信绿，官方/通用推荐"},
    # 简约模板
    {"id": "minimal-gold", "label": "极简金",    "group": "简约模板", "desc": "金色点缀，商务简洁"},
    {"id": "focus-blue",   "label": "专注蓝",    "group": "简约模板", "desc": "蓝色聚焦，适合知识分享"},
    {"id": "elegant-green","label": "优雅绿",    "group": "简约模板", "desc": "绿色雅致，清新正式"},
    {"id": "bold-blue",    "label": "粗体蓝",    "group": "简约模板", "desc": "蓝色粗体标题，醒目"},
]


@app.get("/api/wechat-themes")
def get_wechat_themes():
    """返回可用的公众号主题列表"""
    return WECHAT_THEMES


class RewriteRequest(BaseModel):
    content: str
    mode: str = "口语化"       # 爆款化/口语化/精简/深度展开/加开头钩子/互动结尾
    platform: str = ""         # 抖音脚本/小红书图文/公众号文章/微博段子/通用正文


@app.post("/api/rewrite")
def rewrite_content(req: RewriteRequest):
    """AI 改写：对输入内容按指定模式和平台进行改写，直接返回结果文本"""
    import re

    # ── 模式指令 ──────────────────────────────────────────────
    mode_tasks = {
        "爆款化":    "将以下内容改写为爆款风格：标题吸睛、开头抓人、节奏感强、情绪饱满，适合在社交媒体刷屏传播。",
        "口语化":    "将以下内容改写为轻松口语化风格：去掉书面腔，像和朋友聊天一样自然流畅，有温度感。",
        "精简":      "将以下内容精简压缩：去掉废话和重复表达，保留核心观点，控制在原文 40%-60% 的字数。",
        "深度展开":  "将以下内容深度展开：补充案例、数据、逻辑链，使论点更有说服力和深度，适合长文内容。",
        "加开头钩子": "保持正文不变，在开头加一个强力钩子（疑问/数据冲击/反常识/场景感）吸引读者继续读。",
        "互动结尾":  "保持正文不变，在结尾加一个互动引导结尾（提问/投票/引发讨论），提升评论和互动率。",
    }

    # ── 平台字数要求 ─────────────────────────────────────────
    WORD_MIN = {
        "抖音脚本":   200,
        "小红书图文": 300,
        "公众号文章": 800,
        "微博段子":   100,
        "通用正文":   150,
    }
    platform_hint = {
        "抖音脚本":   "（目标平台：抖音脚本，口语化强，适合配音朗读，每句话简短有力）",
        "小红书图文": "（目标平台：小红书，标题加emoji，段落短，多用感叹号，种草氛围）",
        "公众号文章": "（目标平台：微信公众号，文字更正式，可以稍长，注重排版节奏）",
        "微博段子":   "（目标平台：微博，精炼幽默，100-300字，结尾留金句或反转）",
        "通用正文":   "（通用平台格式，不限制风格）",
    }

    base_task    = mode_tasks.get(req.mode, f"将以下内容进行【{req.mode}】改写。")
    plat_hint    = platform_hint.get(req.platform, f"（目标平台：{req.platform}）" if req.platform else "")
    min_words    = WORD_MIN.get(req.platform, 150)

    # ── Few-Shot 示例库 ──────────────────────────────────────
    # 每个 mode 配 2 个正面示例（输入→输出），展示"好的改写长什么样"
    EXAMPLES = {
        '爆款化': '''
【示例1】
输入：「婚姻需要双方共同努力经营，遇到矛盾时要互相理解包容，这样才能走得更长远。」
输出：「婚姻这东西，最怕的不是吵架，是两个人都憋着不说。我见过太多夫妻，表面上和和气气，背地里各过各的。其实啊，吵一架把话说开，比什么都强。」

【示例2】
输入：「根据研究表明，长期熬夜会导致免疫力下降，对身体健康产生严重影响。」
输出：「熬夜这事，偶尔问题不大，长期真的要命。我不是吓你，是有研究数据撑着的——连续熬夜一周，你身体的免疫力直接拉胯。」''',

        '口语化': '''
【示例1】
输入：「成功的关键在于持续学习和不断实践，只有不断提升自己才能适应环境变化。」
输出：「说白了，你想混得好，就两件事：学和干。不是看书感动自己，是真的下场去练。我见过太多人学了一脑子道理，一做事就抓瞎。」

【示例2】
输入：「该产品具有高效、便捷、环保等多项优势，深受消费者青睐。」
输出：「这东西用起来是真的顺手。我第一次上手的时候就感觉，这设计者是真的懂啊，不像某些产品反人类。」''',

        '精简': '''
【示例1】
输入：「在我们日常生活和工作中，时间管理是一个非常重要的话题，一个人如果能够很好地管理自己的时间，那么他就能够更加高效地完成各种任务，从而有更多的时间去做自己喜欢的事情。」
输出：「会管理时间的人，效率能甩普通人几条街。说白了就是：先想清楚什么重要，什么次要，然后把时间花在刀刃上。」

【示例2】
输入：「在当今社会，互联网技术的发展日新月异，改变了人们的生活方式和工作模式，越来越多的人开始选择远程办公，这种工作方式不仅节省了通勤时间，还提高了工作的灵活性。」
输出：「互联网这东西，彻底改变了我们上班的方式。远程办公越来越常见，省了通勤时间，工作也更灵活了。」''',

        '深度展开': '''
【示例1】
输入：「要学会拒绝别人。」
输出：「不懂拒绝的人，活得有多累？同事找你帮忙，你明明忙得脚不沾地，还是咬牙接下；朋友借钱，你心里一百个不愿意，嘴上还是说"行"。结果呢？自己的事没办好，别人的事也没办好，两头不讨好。学会拒绝不是自私，是对自己负责，也是对别人负责——你帮不了的事硬接，最后搞砸了，对方才真的难受。」

【示例2】
输入：「坚持很重要。」
输出：「为什么有些人做事三分钟热度，有些人能咬牙死磕？我观察下来，差别不在意志力，在于你是不是真的想要。我有个朋友减肥说了三年，每次都失败，不是他不努力，是他其实没那么想瘦——嘴上说说而已。你呢？你想做的事，是真想要，还是只是觉得应该想要？」''',

        '加开头钩子': '''
【示例1】
输入：「职场中要学会与不同性格的人相处...」
输出：「你有没有遇到过那种人？表面上笑眯眯，背后捅刀子。职场三年，我踩过的坑比你听过的道理还多。今天说点真的。」

【示例2】
输入：「健康的饮食习惯对每个人都至关重要...」
输出：「中国有4亿人每天吃的这个，正在慢慢杀死你。不是外卖，是你自己都意识不到的那个习惯。」''',

        '互动结尾': '''
【示例1】
输入：「...以上就是我对这个话题的全部看法」
输出：「...你觉得呢？你有没有类似的经历？欢迎评论区说出来，觉得有用的转给朋友。」

【示例2】
输入：「...这就是我的建议」
输出：「以上，完。觉得有用的话，点个赞；想知道更多，评论区抛出你的问题，我来答。」''',
    }

    few_shot = EXAMPLES.get(req.mode, "")

    # ── AI 禁止词清单（持续补充）───────────────────────────────
    BANNED_WORDS = [
        "此外", "总之", "综上所述", "值得注意的是", "至关重要", "深入探讨",
        "深刻的启示", "引人深省", "不禁让人", "持久的", "格局", "赋能",
        "洞见", "值得深思", "由此可见", "不可磨灭", "见证了", "标志着",
        "体现了", "不言而喻", "毋庸置疑", "而言之", "显然", "无疑",
        "可以看出", "不难发现", "至关重要", "核心", "关键", "本质",
        "层面上", "角度来看", "层面上讲", "从这个意义上",
    ]

    # ── 组装 System Prompt ────────────────────────────────────
    system_prompt = f"""你是一位顶级内容创作者，擅长把素材改写成有灵魂、有温度的真人内容。

【第一步：提炼灵魂——必须先完成】
在改写之前，先从原文中提炼出"3个必须保留的信息点"：
1. 一句话概括核心观点或故事
2. 一个具体细节（数字/名字/场景/案例）
3. 一种核心情绪或态度

格式：【灵魂1】...【灵魂2】...【灵魂3】...

【第二步：执行改写】
{base_task}{plat_hint}

【字数要求】
输出内容必须 ≥ {min_words} 汉字（不含标点符号）。字数不够=任务失败。

【第三步：自检清单】
改写完成后，对照以下清单检查是否通过：
□ 原文3个灵魂信息点是否全部保留？
□ 有没有出现禁用词（此外、总之、关键、核心、赋能、洞见...）？
□ 有没有三段论（首先/其次/最后）？
□ 有没有"这不仅仅是…而是…"结构？
□ 句子长度是否有变化（不是每句都一样长）？
□ 有没有具体数字/名字/场景（不是说"长期"而是说"3年"）？
□ 有没有态度（不是中立报道，而是"我喜欢/我讨厌/我觉得"）？

如果自检不通过，重新改写。

【Few-Shot 参考示例】
{few_shot}

【去AI化规则】

▌词汇红线：以下词出现必须替换
{', '.join(BANNED_WORDS)}
→ 替换示例：此外→还有、总之→说白了、关键→重点、核心→最要紧的、赋能→帮忙、洞见→发现

▌结构红线：
- 禁用三段论（首先→其次→最后）→ 改成自然叙述
- 禁用"这不仅仅是…而是…" → 直接说结论
- 禁用"标志着…" → 删掉，说具体事实
- 禁用金句式收尾 → 换成有态度的话

▌真人写法（给方向，不是禁止清单）：
✅ 句子长短交错——短句打节奏，长句偶尔来一句
✅ 有你明确的态度——喜欢就是喜欢，不爽就是不爽
✅ 说具体的——"3年"好过"长期"，"第一个月"好过"初期"
✅ 允许有点毛边——真人说话就是有点散，不是每句都对仗
✅ 像给朋友发消息那样写——自然、亲切、有画面感

直接输出【改写结果】，不要输出自检过程、不要前言后记。"""

    user_prompt = req.content

    try:
        # 先估算输入字数，用于判断"精简"模式的目标
        input_len = len(re.sub(r'\s+', '', req.content))
        if req.mode == "精简":
            user_prompt = f"【原文约{input_len}字，目标输出{int(input_len * 0.4)}-{int(input_len * 0.6)}字】\n\n" + req.content

        result, model_used = ai_extract(system_prompt, user_prompt, max_tokens=8000, temperature=0.85)
        return {"result": result, "model": model_used}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI改写失败：{str(e)}")


# ═══════════════════════════════════════════════════════════════════════════
# P2: 多版本 Ensemble + 4-Pass Humanization
# ═══════════════════════════════════════════════════════════════════════════

def _humanize_pass(text: str) -> str:
    """4-Pass Humanization 后处理"""
    import re

    # Pass 1: 移除 AI 高频词（扩展清单）
    AI_WORDS = [
        '此外', '总之', '综上所述', '值得注意的是', '至关重要', '深入探讨',
        '深刻的启示', '引人深省', '不禁让人', '持久的', '格局', '赋能',
        '洞见', '值得深思', '由此可见', '不可磨灭', '见证了', '标志着',
        '体现了', '不言而喻', '毋庸置疑', '而言之', '显然', '无疑',
        '可以看出', '不难发现', '核心', '关键', '本质', '层面上',
        '角度来看', '层面上讲', '从这个意义上', '值得注意的是',
        '值得注意的是', '不难发现', '由此可见', '可以说',
    ]
    for w in AI_WORDS:
        # 简单替换为口语化表达
        replacements = {
            '此外': '还有', '总之': '说白了', '综上所述': '归根结底',
            '值得注意的是': '要我说', '至关重要': '特别重要', '深入探讨': '细说',
            '深刻的启示': '挺有启发的', '引人深省': '值得想想', '不禁让人': '让人',
            '持久的': '长期的', '格局': '眼界', '赋能': '帮忙', '洞见': '发现',
            '值得深思': '值得想想', '由此可见': '可见', '不可磨灭': '不会消失的',
            '见证了': '看到了', '标志着': '说明', '体现了': '反映了',
            '不言而喻': '不用说', '毋庸置疑': '不用怀疑', '而言之': '总之',
            '显然': '很明显', '无疑': '不用说', '可以看出': '可见',
            '不难发现': '容易看到', '核心': '最核心', '关键': '最关键',
            '本质': '最本质', '层面上': '方面',
        }
        text = text.replace(w, replacements.get(w, '很'))

    # Pass 2: 打破均匀句长——在连续3个短句后插入一个较长句子
    sentences = re.split(r'([。！？；\n])', text)
    result = []
    short_count = 0
    for i in range(0, len(sentences), 2):
        s = sentences[i]
        punct = sentences[i+1] if i+1 < len(sentences) else ''
        if not s.strip():
            continue
        word_count = len(s.strip())
        if word_count < 15:
            short_count += 1
        else:
            short_count = 0
        result.append(s + punct)
        # 连续短句过多时，强制合并
        if short_count >= 3 and i+2 < len(sentences):
            next_s = sentences[i+2].strip() if i+2 < len(sentences) else ''
            if next_s:
                result[-1] = result[-1].rstrip('。！？；') + '，' + next_s
                short_count = 0

    text = ''.join(result)

    # Pass 3: 移除"可引用金句结尾"——如果结尾像格言，加一句口语化吐槽
    text = text.strip()
    if text and len(text) > 20:
        last_sentence = re.split(r'[。！？；\n]', text)[-2] if len(re.split(r'[。！？；\n]', text)) > 1 else ''
        if last_sentence and (len(last_sentence) < 30 and ('。' not in last_sentence[-10:])):
            # 结尾像金句，加一句口语化的话
            text = text.rstrip('。！？') + '。说完了，就这样。'

    # Pass 4: 注入自我引用（只在合适位置加一句）
    if '我认为' not in text and '我觉得' not in text and len(text) > 100:
        # 在中间位置插入
        mid = len(text) // 2
        text = text[:mid] + '我自己的想法是，' + text[mid:]

    return text


class RewriteEnsembleRequest(BaseModel):
    content: str
    mode: str = "口语化"
    platform: str = ""


@app.post("/api/rewrite-ensemble")
def rewrite_ensemble(req: RewriteEnsembleRequest):
    """多版本改写——用两种不同风格生成2个版本，让用户选择更好的"""
    # 复用 rewrite_content 的逻辑，但用不同 temperature
    # 版本1: 温度 0.7（保守、精准）
    # 版本2: 温度 0.95（创意、多变）
    import re as _re2

    mode_tasks = {
        "爆款化":    "将以下内容改写为爆款风格：标题吸睛、开头抓人、节奏感强、情绪饱满，适合在社交媒体刷屏传播。",
        "口语化":    "将以下内容改写为轻松口语化风格：去掉书面腔，像和朋友聊天一样自然流畅，有温度感。",
        "精简":      "将以下内容精简压缩：去掉废话和重复表达，保留核心观点，控制在原文 40%-60% 的字数。",
        "深度展开":  "将以下内容深度展开：补充案例、数据、逻辑链，使论点更有说服力和深度，适合长文内容。",
        "加开头钩子": "保持正文不变，在开头加一个强力钩子（疑问/数据冲击/反常识/场景感）吸引读者继续读。",
        "互动结尾":  "保持正文不变，在结尾加一个互动引导结尾（提问/投票/引发讨论），提升评论和互动率。",
    }
    WORD_MIN = {
        "抖音脚本":   200, "小红书图文": 300, "公众号文章": 800, "微博段子": 100, "通用正文": 150,
    }
    platform_hint = {
        "抖音脚本":   "（目标平台：抖音脚本，口语化强，适合配音朗读，每句话简短有力）",
        "小红书图文": "（目标平台：小红书，标题加emoji，段落短，多用感叹号，种草氛围）",
        "公众号文章": "（目标平台：微信公众号，文字更正式，可以稍长，注重排版节奏）",
        "微博段子":   "（目标平台：微博，精炼幽默，100-300字，结尾留金句或反转）",
        "通用正文":   "（通用平台格式，不限制风格）",
    }

    base_task = mode_tasks.get(req.mode, f"将以下内容进行【{req.mode}】改写。")
    plat_hint = platform_hint.get(req.platform, f"（目标平台：{req.platform}）" if req.platform else "")
    min_words = WORD_MIN.get(req.platform, 150)

    system_base = f"""你是一位顶级内容创作者，擅长把素材改写成有灵魂、有温度的真人内容。

【本次任务】
{base_task}{plat_hint}
【输出字数要求】≥ {min_words} 汉字。
【去AI化】禁用词：此外、总之、关键、核心、赋能、洞见、值得注意的是、深入探讨。替换：此外→还有、总之→说白了。
【真人写法】✅ 句子长短交错 ✅ 有态度 ✅ 说具体的（"3年"好过"长期"）✅ 像给朋友发消息
直接输出改写结果，不要前言后记。"""

    results = []
    # 跑两个版本：低温（0.7）和高温（0.95）
    for temp, label in [(0.7, "稳重版"), (0.95, "创意版")]:
        try:
            text, model = ai_extract(system_base, req.content, max_tokens=8000, temperature=temp)
            # 应用 humanization 后处理
            text = _humanize_pass(text)
            char_count = len(_re2.sub(r'\s+', '', text))
            results.append({
                "label": label,
                "text": text,
                "chars": char_count,
                "model": model,
                "temp": temp,
            })
        except Exception as e:
            results.append({"label": label, "error": str(e), "text": ""})

    return {"versions": results}


class WechatFormatRequest(BaseModel):
    content: str
    theme: str = "wechat-native"


@app.post("/api/wechat-format")
def wechat_format(req: WechatFormatRequest):
    """将 Markdown 排版为微信公众号兼容的内联样式 HTML"""
    import sys
    import re as _re
    from pathlib import Path

    format_dir = BASE_DIR / "wechat-format"
    if str(format_dir) not in sys.path:
        sys.path.insert(0, str(format_dir))

    try:
        import format as wfmt
    except ImportError as e:
        raise HTTPException(500, f"排版模块加载失败：{e}")

    # 验证主题存在
    valid_ids = {t["id"] for t in WECHAT_THEMES}
    theme_name = req.theme if req.theme in valid_ids else "wechat-native"

    try:
        theme = wfmt.load_theme(theme_name)

        # 直接执行核心处理流程，绕开 vault_root os.walk（Windows 根目录遍历会卡死）
        content = req.content
        input_path = Path("article.md")
        output_dir = BASE_DIR / "data" / "wechat-output"
        output_dir.mkdir(parents=True, exist_ok=True)

        # 通用预处理（与 format_for_output 一致，跳过 wikilinks/本地图片）
        title = wfmt.extract_title(content, input_path)
        word_count = wfmt.count_words(content)
        content = wfmt.strip_frontmatter(content)
        content = wfmt.fix_cjk_spacing(content)
        content = wfmt.fix_cjk_bold_punctuation(content)
        content = wfmt.process_callouts(content)
        content = wfmt.process_manual_footnotes(content)
        content = wfmt.process_fenced_containers(content)
        content = _re.sub(r'~~(.+?)~~', r'<del>\1</del>', content)
        # 跳过 convert_wikilinks（需要遍历本地文件系统，Windows 上 vault_root='/' 会崩溃）
        # 跳过 copy_markdown_images（用户粘贴的是纯文字，无本地图片）

        html = wfmt.md_to_html(content)
        html, footnote_html = wfmt.extract_links_as_footnotes(html)
        html = wfmt.inject_inline_styles(html, theme)
        if footnote_html:
            footnote_html = wfmt.inject_inline_styles(footnote_html, theme, skip_wrapper=True)
        html = wfmt.convert_image_captions(html)

        full_html = html + ("\n" + footnote_html if footnote_html else "")
        return {
            "html": full_html,
            "title": title,
            "word_count": word_count,
            "theme": theme_name,
        }
    except Exception as e:
        import traceback
        raise HTTPException(500, f"排版失败：{e}\n{traceback.format_exc()}")


@app.post("/api/wechat-preview")
def wechat_preview(req: WechatFormatRequest):
    """返回完整可渲染的 HTML 页面（旧接口保留兼容）"""
    result = wechat_format(req)
    body_html = result["html"]
    full_page = _build_preview_page(body_html)
    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=full_page, media_type="text/html; charset=utf-8")


# ─── 主题画廊（Gallery）─────────────────────────────────────────────────
GALLERY_THEMES = [
    # 深度长文（4）
    "newspaper", "magazine", "ink", "coffee-house",
    # 科技产品（4）
    "bytedance", "github", "sspai", "midnight",
    # 文艺随笔（4）
    "terracotta", "mint-fresh", "sunset-amber", "lavender-dream",
    # 活力动态（4）
    "sports", "bauhaus", "chinese", "wechat-native",
]

GALLERY_DEMO_MARKDOWN = """\
## 主要功能

在数字化时代，**内容创作**变得越来越重要。一款好的排版工具，能让你的文章在众多内容中**脱颖而出**。

> 好的排版不只是视觉享受，更是对读者的尊重。

### 核心亮点

- 完整的 Markdown 语法支持
- 精美的主题样式
- 一键复制到微信发布

1. 撰写你的内容
2. 选择喜欢的风格
3. 一键复制粘贴

---

### 代码示例

`inline code` 也是支持的。

```python
def hello():
    print("Hello, World!")
```

| 功能 | 状态 |
|------|------|
| 实时预览 | 已支持 |
| 主题选择 | 已支持 |

> [!tip] 小技巧
> 选择适合你文章风格的主题，效果更佳。
"""


def _render_gallery_theme(tid: str, theme_data: dict, html: str, footnote_html: str) -> tuple:
    """渲染单个主题（用于并行 gallery）"""
    # 每个线程需要自己的 import，因为 sys.path 是线程独立的
    import sys
    from pathlib import Path
    format_dir = Path(__file__).parent.parent / "wechat-format"
    if str(format_dir) not in sys.path:
        sys.path.insert(0, str(format_dir))
    import format as wfmt
    rendered = wfmt.inject_inline_styles(html, theme_data)
    fn_rendered = ""
    if footnote_html:
        fn_rendered = wfmt.inject_inline_styles(footnote_html, theme_data, skip_wrapper=True)
    return tid, rendered + ("\n" + fn_rendered if fn_rendered else "")


def generate_gallery_html(rendered_map: dict, theme_map: dict,
                         theme_ids: list, title: str, word_count: int,
                         recommended: list = None) -> str:
    """生成主题画廊页面 HTML"""
    if recommended is None:
        recommended = []

    base_style = (
        "*{box-sizing:border-box}"
        "html,body{margin:0;padding:0;background:#fff}"
        "body{padding:16px 24px;font-family:-apple-system,'PingFang SC','Microsoft YaHei',sans-serif;"
        "font-size:15px;line-height:1.75;color:#333}"
        "img{max-width:100%;display:block;margin:8px auto}"
    )

    GROUPS = [
        ("深度长文", ["newspaper", "magazine", "ink", "coffee-house"]),
        ("科技产品", ["bytedance", "github", "sspai", "midnight"]),
        ("文艺随笔", ["terracotta", "mint-fresh", "sunset-amber", "lavender-dream"]),
        ("活力动态", ["sports", "bauhaus", "chinese", "wechat-native"]),
    ]

    buttons_html = ""
    btn_index = 0
    for group_name, group_ids in GROUPS:
        group_tids = [t for t in group_ids if t in theme_ids]
        if not group_tids:
            continue
        buttons_html += f'<div class="theme-group"><span class="group-label">{group_name}</span>'
        for tid in group_tids:
            theme = theme_map[tid]
            accent = theme.get("colors", {}).get("accent", "#333")
            active = " active" if btn_index == 0 else ""
            is_recommended = " recommended" if tid in recommended else ""
            name = theme.get("name", tid)
            rec_label = '<span class="rec-badge">推荐</span>' if tid in recommended else ""
            buttons_html += (
                f'<button class="theme-btn{active}{is_recommended}" data-theme="{tid}" '
                f'onclick="switchTheme(\'{tid}\')">'
                f'<span class="theme-dot" style="background:{accent}"></span>'
                f'{name}{rec_label}</button>'
            )
            btn_index += 1
        buttons_html += '</div>\n'

    previews_html = ""
    for i, tid in enumerate(theme_ids):
        display = "block" if i == 0 else "none"
        previews_html += (
            f'<div class="theme-preview" data-theme="{tid}" '
            f'style="display:{display}">{rendered_map[tid]}</div>\n'
        )

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{title} - 主题选择</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif; background: #f5f5f7; min-height: 100vh; }}
.toolbar {{ position: fixed; top: 0; left: 0; right: 0; background: rgba(255,255,255,0.92); backdrop-filter: blur(20px); -webkit-backdrop-filter: blur(20px); border-bottom: 1px solid #e0e0e0; padding: 14px 24px; display: flex; align-items: center; justify-content: space-between; z-index: 200; box-shadow: 0 1px 4px rgba(0,0,0,0.06); }}
.toolbar-left {{ display: flex; align-items: center; gap: 16px; }}
.toolbar-title {{ font-size: 15px; font-weight: 600; color: #1d1d1f; }}
.toolbar-meta {{ font-size: 13px; color: #86868b; }}
.toolbar-hint {{ font-size: 13px; color: #07c160; font-weight: 500; }}
.main-container {{ max-width: 700px; margin: 80px auto 40px; padding: 0 20px; display: flex; flex-direction: column; align-items: center; gap: 24px; }}
.theme-buttons {{ display: flex; flex-wrap: wrap; justify-content: center; gap: 10px; max-width: 600px; }}
.theme-group {{ display: flex; flex-wrap: wrap; gap: 8px; align-items: center; margin-bottom: 8px; }}
.group-label {{ font-size: 12px; color: #86868b; font-weight: 600; min-width: 60px; letter-spacing: 0.5px; }}
.theme-btn {{ display: inline-flex; align-items: center; gap: 6px; padding: 6px 14px; background: #fff; border: 2px solid #e0e0e0; border-radius: 20px; font-size: 13px; font-weight: 500; color: #333; cursor: pointer; transition: all 0.2s; font-family: inherit; }}
.theme-btn:hover {{ border-color: #ccc; background: #fafafa; }}
.theme-btn.active {{ border-color: #07c160; background: rgba(7,193,96,0.06); color: #07c160; }}
.theme-btn.recommended {{ position: relative; }}
.theme-btn.recommended .rec-badge {{ position: absolute; top: -8px; right: -6px; background: #ff9500; color: #fff; font-size: 10px; line-height: 1; padding: 2px 5px; border-radius: 6px; font-weight: 600; pointer-events: none; }}
.theme-dot {{ display: inline-block; width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }}
.phone-frame {{ width: 100%; max-width: 415px; background: #fff; border-radius: 16px; box-shadow: 0 8px 40px rgba(0,0,0,0.12); overflow: hidden; }}
.phone-header {{ background: #ededed; padding: 12px 16px; display: flex; align-items: center; justify-content: center; font-size: 13px; color: #999; border-bottom: 1px solid #e0e0e0; }}
.preview-scroll {{ max-height: 600px; overflow-y: auto; padding: 20px; }}
.theme-preview {{ font-family: -apple-system, "PingFang SC", sans-serif; }}
.action-btn {{ width: 100%; max-width: 415px; padding: 14px 0; background: #07c160; color: #fff; border: none; border-radius: 10px; font-size: 16px; font-weight: 600; cursor: pointer; transition: all 0.2s; letter-spacing: 0.5px; font-family: inherit; }}
.action-btn:hover {{ background: #06ae56; }}
.action-btn:active {{ transform: scale(0.99); }}
.action-btn.copied {{ background: #34c759; }}
.font-size-selector {{ display: flex; align-items: center; justify-content: center; gap: 8px; margin-bottom: 16px; padding: 8px 0; }}
.fs-label {{ font-size: 13px; color: #999; }}
.fs-btn {{ padding: 4px 12px; border: 1px solid #ddd; border-radius: 16px; background: #fff; font-size: 13px; color: #666; cursor: pointer; transition: all 0.2s; font-family: inherit; }}
.fs-btn:hover {{ border-color: #07C160; color: #07C160; }}
.fs-btn.active {{ background: #07C160; border-color: #07C160; color: #fff; }}
@media (max-width: 600px) {{ .toolbar {{ padding: 10px 14px; }} .toolbar-hint {{ display: none; }} .main-container {{ margin-top: 68px; padding: 0 12px; }} .theme-buttons {{ display: grid; grid-template-columns: 1fr 1fr; width: 100%; }} .theme-btn {{ justify-content: center; }} .phone-frame {{ border-radius: 0; box-shadow: none; }} .preview-scroll {{ max-height: 500px; }} }}
</style>
</head>
<body>
<div class="toolbar">
    <div class="toolbar-left">
        <span class="toolbar-title">{title}</span>
        <span class="toolbar-meta">{word_count} 字</span>
    </div>
    <span class="toolbar-hint">选一个你喜欢的风格</span>
</div>
<div class="main-container">
    <div class="font-size-selector">
        <span class="fs-label">字号</span>
        <button class="fs-btn active" data-size="15" onclick="switchFontSize(15)">15px</button>
        <button class="fs-btn" data-size="16" onclick="switchFontSize(16)">16px</button>
    </div>
    <div class="theme-buttons">
{buttons_html}
    </div>
    <div class="phone-frame">
        <div class="phone-header">微信公众号预览</div>
        <div class="preview-scroll" id="previewScroll">
{previews_html}
        </div>
    </div>
    <button class="action-btn" id="copyBtn" onclick="applyTheme()">复制内容到剪贴板 → 公众号粘贴</button>
</div>
<script>
var selectedTheme = '{theme_ids[0] if theme_ids else ""}';
function switchTheme(themeId) {{
    var previews = document.querySelectorAll('.theme-preview');
    for (var i = 0; i < previews.length; i++) {{ previews[i].style.display = 'none'; }}
    var target = document.querySelector('.theme-preview[data-theme="' + themeId + '"]');
    if (target) {{ target.style.display = 'block'; }}
    var btns = document.querySelectorAll('.theme-btn');
    for (var i = 0; i < btns.length; i++) {{
        if (btns[i].getAttribute('data-theme') === themeId) {{ btns[i].classList.add('active'); }} else {{ btns[i].classList.remove('active'); }}
    }}
    selectedTheme = themeId;
    document.getElementById('previewScroll').scrollTop = 0;
}}
function applyTheme() {{
    var target = document.querySelector('.theme-preview[data-theme="' + selectedTheme + '"]');
    if (!target) return;
    var range = document.createRange();
    range.selectNodeContents(target);
    var sel = window.getSelection();
    sel.removeAllRanges();
    sel.addRange(range);
    try {{ document.execCommand('copy'); }} catch(e) {{}}
    sel.removeAllRanges();
    var btn = document.getElementById('copyBtn');
    var activeBtn = document.querySelector('.theme-btn.active');
    var themeName = activeBtn ? activeBtn.textContent.trim().replace('推荐','').trim() : selectedTheme;
    btn.textContent = '已复制「' + themeName + '」 ✓';
    btn.classList.add('copied');
    setTimeout(function() {{ btn.textContent = '复制内容到剪贴板 → 公众号粘贴'; btn.classList.remove('copied'); }}, 3000);
}}
function switchFontSize(size) {{
    document.querySelectorAll('.fs-btn').forEach(function(btn) {{
        if (parseInt(btn.dataset.size) === size) {{ btn.classList.add('active'); }} else {{ btn.classList.remove('active'); }}
    }});
    document.querySelectorAll('.theme-preview').forEach(function(preview) {{
        preview.querySelectorAll('p, span, section').forEach(function(el) {{
            var fs = parseInt(el.style.fontSize);
            if (fs >= 14 && fs <= 18) {{ el.style.fontSize = size + 'px'; }}
        }});
    }});
}}
document.addEventListener('DOMContentLoaded', function() {{}});
</script>
</body>
</html>"""


@app.post("/api/wechat-gallery")
def wechat_gallery(req: WechatFormatRequest):
    """返回主题画廊页面，可一次预览所有主题"""
    import re as _re
    from pathlib import Path
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from fastapi.responses import HTMLResponse

    format_dir = BASE_DIR / "wechat-format"
    if str(format_dir) not in sys.path:
        sys.path.insert(0, str(format_dir))

    try:
        import format as wfmt
    except ImportError as e:
        raise HTTPException(500, f"排版模块加载失败：{e}")

    content = req.content or GALLERY_DEMO_MARKDOWN

    # 预处理（与 wechat_format 一致，跳过 wikilinks/本地图片）
    title = wfmt.extract_title(content, Path("article.md"))
    word_count = wfmt.count_words(content)
    content = wfmt.strip_frontmatter(content)
    content = wfmt.fix_cjk_spacing(content)
    content = wfmt.fix_cjk_bold_punctuation(content)
    content = wfmt.process_callouts(content)
    content = wfmt.process_manual_footnotes(content)
    content = wfmt.process_fenced_containers(content)
    content = _re.sub(r'~~(.+?)~~', r'<del>\1</del>', content)

    html = wfmt.md_to_html(content)
    html, footnote_html = wfmt.extract_links_as_footnotes(html)

    # 加载所有画廊主题
    theme_map = {}
    gallery_theme_ids = []
    for tid in GALLERY_THEMES:
        try:
            theme_map[tid] = wfmt.load_theme(tid)
            gallery_theme_ids.append(tid)
        except Exception:
            pass

    if not gallery_theme_ids:
        return HTMLResponse("<html><body><p>无可用主题</p></body></html>", media_type="text/html; charset=utf-8")

    # 并行渲染所有主题
    rendered_map = {}
    with ThreadPoolExecutor(max_workers=min(8, len(gallery_theme_ids))) as executor:
        futures = {
            executor.submit(_render_gallery_theme, tid, theme_map[tid], html, footnote_html): tid
            for tid in gallery_theme_ids
        }
        for future in as_completed(futures):
            tid, rendered = future.result()
            rendered_map[tid] = rendered

    gallery_html = generate_gallery_html(
        rendered_map, theme_map, gallery_theme_ids,
        title, word_count, recommended=[req.theme] if req.theme in gallery_theme_ids else []
    )

    return HTMLResponse(content=gallery_html, media_type="text/html; charset=utf-8")


# ─── 预览缓存（token → HTML）─────────────────────────────────────────────
import uuid as _uuid
_preview_cache: dict = {}  # token -> full_html

def _build_preview_page(body_html: str) -> str:
    base_style = (
        "*{box-sizing:border-box}"
        "html,body{margin:0;padding:0;background:#fff}"
        "body{padding:16px 24px;font-family:-apple-system,'PingFang SC','Microsoft YaHei',sans-serif;"
        "font-size:15px;line-height:1.75;color:#333}"
        "img{max-width:100%;display:block;margin:8px auto}"
    )
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>{base_style}</style>
</head>
<body>{body_html}</body>
</html>"""


class PreviewStoreRequest(BaseModel):
    content: str
    theme: str = "default"


@app.post("/api/preview-store")
def preview_store(req: PreviewStoreRequest):
    """生成排版并缓存到内存，返回 token；前端用 GET /api/preview/{token} 在 iframe 中展示"""
    fmt_result = wechat_format(WechatFormatRequest(content=req.content, theme=req.theme))
    token = _uuid.uuid4().hex
    _preview_cache[token] = _build_preview_page(fmt_result["html"])
    # 最多保留 20 条缓存，防止内存膨胀
    if len(_preview_cache) > 20:
        oldest = next(iter(_preview_cache))
        del _preview_cache[oldest]
    return {
        "token": token,
        "html": fmt_result["html"],       # 仍返回 html 供复制
        "word_count": fmt_result.get("word_count"),
        "theme": fmt_result.get("theme", req.theme),
    }


@app.get("/api/preview/{token}")
def preview_get(token: str):
    """GET 接口，返回缓存的完整 HTML 预览页面；可直接作为 iframe src"""
    from fastapi.responses import HTMLResponse
    html = _preview_cache.get(token)
    if not html:
        return HTMLResponse("<html><body><p style='color:#999;text-align:center;padding:40px'>预览已过期，请重新生成排版</p></body></html>")
    return HTMLResponse(content=html, media_type="text/html; charset=utf-8")





# ─── 朴树之道分享 ──────────────────────────────────────────────────────

PUSHUTREE_SYSTEM_PROMPT = """你是一个真正读过无数书的人，现在做书籍推荐视频/文案。

【你的核心定位】
你不是在写读书笔记，不是在写书评，也不是在讲故事。
你是在做"推荐词"——你要让一个完全不了解这本书的观众，在看完你这期内容之后：
1. 得到一个对他生活真正有用的认知/方法/感悟
2. 心里有点被触动（情绪共鸣）
3. 产生想去读这本书的冲动

每一期内容都是独立完整的。观众不需要看前一期，也不需要看后一期。
这不是连续剧，是系列推荐词——每期讲书里一个核心点，从不同角度打动人。

【什么是好的推荐词——学董宇辉/朴树之道】

董宇辉推荐书的方式：
- 先说一个让人心里一颤的问题或生活场景（开头抓人）
- 然后说书里怎么回答这个问题，或者这本书给了他什么新的认知
- 说自己的感受，不是复述书的内容
- 金句多，但不是鸡汤，是真实的洞察
- 让人觉得：这个人是真的读过，而且真的有感悟

朴树之道的感觉：
- 不装，不端，真实
- 说话像跟朋友聊天，不像在开讲座
- 有自己的判断和态度，不是中立的介绍

【硬性禁止】
❌ 叙事型写法："这本书讲了一个故事……""在书的第X章，作者提到……"
❌ 读书报告结构："这本书共分三部分，首先……其次……最后……"
❌ AI废话：深刻、启示、洞见、值得深思、引人深省、综上所述、由此可见
❌ 连续剧式开头："上一期我们讲到……""今天我们继续……"
❌ 讲述书的情节，像在讲故事

【每期推荐词的正确结构】
① 开头钩子（1-3句）：一个让人停下来的问题、反常识判断、或真实生活场景
   - 要打中观众的某个真实困惑或痛点
   - 前三句决定观众走不走

② 核心认知（2-3段）：
   - 书里给出的答案/方法/视角——但用你自己的话说
   - 要具体，用书里真实的案例/数据/故事支撑
   - 让人有"原来如此"的感觉

③ 为什么你/观众需要这本书（1-2段）：
   - 这个认知能帮观众解决什么真实问题
   - 读完这本书，观众的某一件事会变得不同

④ 结尾推荐句（1-2句）：
   - 不是"这本书值得一读"这种废话
   - 是一句让人想截图或转发的话
   - 带情绪，带态度，真实

【字数要求】：600-900字，干货密度高，不要凑字数

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【去AI化——写作灵魂规则，让文字像真人说出来的】
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

这是口播稿，更不能有AI味。观众一眼就能听出来那种"机器人感"。

▌绝对禁止的AI词汇：
此外、总之、综上所述、值得注意的是、至关重要、深入探讨、深刻的启示、引人深省、
不禁让人、持久的影响、格局、赋能、洞见、值得深思、由此可见、不可磨灭、
见证了、标志着、体现了、不言而喻、毋庸置疑、可以说、某种程度上

▌禁止的AI写作结构（口播特别注意）：
- 三段论公式开头："有三个方面……" → 直接讲第一个，自然带出下一个
- 否定排比："这不仅仅是……而是……" → 直接说结论
- 宣传性形容词："深刻的""颠覆性的""令人震撼的" → 删掉，让事实本身说话
- 过长的句子（超过25字）→ 拆成两句，口播要短句

▌真人口播的写法：
- 多用"你"，少用"我们"——直接跟观众说话
- 允许说"说实话""我觉得""坦白说"——口语感
- 举的例子要具体，有数字有细节（"三年""每天早上6点""第87页"比"长期""坚持""书中"更真实）
- 句子有轻有重，有长有短——模拟真人说话的节奏
- 偶尔留一个问句让观众心里有点颤——不需要他回答，但他会停下来想"""


def _fix_json_bare_quotes(json_str: str) -> str:
    """修复 JSON 字符串值内部的裸双引号（AI 常见输出错误）。
    
    逐字符扫描：在字符串值内遇到未转义的双引号时替换为单引号。
    这样 {"key": "he said "hello""} 会被修复为 {"key": "he said 'hello'"}
    """
    result = []
    in_string = False
    i = 0
    prev_was_key = False  # 上一个字符串是键（不在值里）
    
    while i < len(json_str):
        c = json_str[i]
        
        if not in_string:
            if c == '"':
                in_string = True
                result.append(c)
            else:
                result.append(c)
        else:
            if c == '\\' and i + 1 < len(json_str):
                # 合法转义序列，原样保留
                result.append(c)
                result.append(json_str[i + 1])
                i += 2
                continue
            elif c == '"':
                # 字符串结束，检查下一个非空白字符是否为 JSON 合法后续符
                j = i + 1
                while j < len(json_str) and json_str[j] in ' \t\r\n':
                    j += 1
                next_c = json_str[j] if j < len(json_str) else ''
                if next_c in ',:}]':
                    # 这是合法的字符串结束引号
                    in_string = False
                    result.append(c)
                else:
                    # 这是字符串内部的裸引号，替换为单引号
                    result.append("'")
            else:
                result.append(c)
        i += 1
    
    return ''.join(result)


def pushutree_plan(book_name: str, text: str, episode_count: int, style: str) -> list[dict]:
    """第一步：策划——分析全书找痛点，生成标题+大纲"""
    user_prompt = f"""请仔细阅读《{book_name}》的内容，用你作为一个真正读过这本书的人的眼光，找出{episode_count}个最值得分享的核心主题。

不要找那种"第X章讲了什么"的表面总结，要找：
- 让你看完会停下来想一想的观点
- 打破常见认知误区的内容
- 能帮普通人解决真实生活困惑的道理
- 书里藏着但很少人注意到的深层规律

【书籍内容】
{text[:60000]}

【输出要求】
请严格按照以下JSON格式输出，不要输出其他任何内容：
{{
  "episodes": [
    {{
      "ep_no": 1,
      "title": "标题（15字以内，有反差感或冲击力，不用你不知道的XX这类老套标题）",
      "angle": "这期的切入角度（一句话，说清楚从哪个生活场景或困惑切入）",
      "core_insight": "这期最核心的认知或规律（不是金句，是一个真实的洞察）",
      "pain_point": "击中的读者真实痛点（要具体，不要写迷茫这种大词）",
      "source_hint": "书中哪个章节或故事或案例或数据可以支撑",
      "why_useful": "读者看完能学到什么，能用在哪里（一句话）"
    }}
  ]
}}

重要：JSON字段的值必须是合法字符串，不能包含未转义的双引号（如需引用，请用书名号《》或单引号'代替）。
风格要求：{style}
共需{episode_count}期，每期角度不重复，要覆盖全书的精华，不要只集中在前几章。"""

    raw, _model = ai_extract(
        system_prompt=PUSHUTREE_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        max_tokens=4000,
        temperature=0.7,
    )
    raw = raw.strip()
    # 提取JSON（容错：处理AI在JSON前后输出废话、markdown代码块、中文引号等）
    import re
    # 去掉 markdown 代码块包裹 ```json ... ```
    raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
    raw = re.sub(r'\s*```\s*$', '', raw, flags=re.MULTILINE)
    # 替换中文引号为英文引号（常见AI输出错误）
    raw = raw.replace('\u201c', '"').replace('\u201d', '"').replace('\u2018', "'").replace('\u2019', "'")
    # 提取最外层 JSON 对象
    m = re.search(r'\{[\s\S]*\}', raw)
    if m:
        json_str = m.group()
        # 第一次尝试：直接解析
        try:
            data = json.loads(json_str)
            return data.get("episodes", [])
        except json.JSONDecodeError as je:
            print(f"[pushutree_plan] 第一次JSON解析失败: {je}，尝试修复裸引号...")
        # 第二次尝试：修复字段值内的裸双引号
        # 思路：逐字扫描，在字符串值内部遇到裸双引号时替换为单引号
        try:
            fixed = _fix_json_bare_quotes(json_str)
            data = json.loads(fixed)
            print(f"[pushutree_plan] 修复裸引号后解析成功")
            return data.get("episodes", [])
        except Exception as je2:
            print(f"[pushutree_plan] 修复后仍失败: {je2}\n原始内容(前800): {json_str[:800]}")
            return []
    print(f"[pushutree_plan] 未找到JSON结构，原始内容(前500): {raw[:500]}")
    return []


def pushutree_write(book_name: str, text: str, episode: dict, style: str) -> str:
    """第二步：撰写——写一篇独立的书籍推荐词"""
    user_prompt = f"""现在为《{book_name}》写第{episode['ep_no']}期书籍推荐词。

【本期要传递的核心内容】
- 这期聚焦的角度：{episode['angle']}
- 核心认知：{episode.get('core_insight', episode.get('core_quote',''))}
- 观众痛点：{episode['pain_point']}
- 书中支撑：{episode['source_hint']}
- 观众能得到什么：{episode.get('why_useful','一个对生活有用的新认知')}

【书籍原文（供参考提取真实内容）】
{text[:50000]}

---

【写作任务】
写一篇让观众不想划走、看完有收获、想去读这本书的推荐词。

这不是读书笔记，不是书评，不是连续剧的某一集。
这是一篇独立完整的推荐词，观众看了这一期，不需要看其他期。

【必须做到的三件事】

第一：开头要抓人（前3句）
不要从"这本书讲了……"开始。
要从观众的真实感受/困惑/生活场景开始。
比如：
- 直接抛出一个让人心里一紧的问题
- 说一个很多人有但没说出来的感受
- 说一个和常识相反的判断

第二：干货要真实（中间部分）
用书里真实的案例、数据、观点来说话。
不要只说结论，要说"书里是怎么说的"——但用你自己的话，不是复制粘贴。
每个认知点要对应一个"观众能用在哪里"。

第三：结尾要有力量
不是"这本书值得一读"。
是一句让人想截图的话——有点哲学，有点反常识，或者说出了很多人想说又说不出来的话。

【风格】：{style}
【字数】：600-900字
【格式】：自然段落，口语化，短句为主，不用序号标题

【去AI化——写的时候就不能有AI味，不要等到精修再改】
✅ 用"你"开头或在关键句里直接和观众说话
✅ 举具体例子（书里的真实数字、细节、场景），不说空话
✅ 句子长短交错——3-5个短句打节奏，偶尔一句长句讲道理
✅ 有自己的态度——"我觉得这书有一点特别重要""说实话这个观点很少有人能做到"
❌ 绝对不写：此外、总之、由此可见、深刻的、至关重要、值得注意的是
❌ 绝对不用三段论公式：首先……其次……最后……"""

    draft, _model = ai_extract(
        system_prompt=PUSHUTREE_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        max_tokens=3000,
        temperature=0.85,
    )
    return draft.strip()


def pushutree_polish(book_name: str, draft: str, episode: dict) -> str:
    """第三步：精修——观众视角终审 + 去AI化深度清洗"""
    user_prompt = f"""对下面这篇《{book_name}》书籍推荐词做最终精修。

【精修标准——你是一个极度挑剔的观众】

先以观众身份看一遍，问自己：

❶ 开头三句，我会不会划走？
   - 如果开头是"这本书讲了……" → 必须重写，从观众痛点切入
   - 如果开头是废话过渡 → 删掉，直接进正文
   
❷ 中间内容，我学到了什么？
   - 有没有一个对我日常生活真正有用的认知？
   - 说的是书里真实的东西，还是泛泛而谈？

❸ 这像推荐词还是读书报告？
   - 推荐词：说人话，有态度，让我想读这本书
   - 读书报告：复述内容，中立无感，让我想睡觉
   - 如果是读书报告味道 → 砍掉复述段落，换成"这对你意味着什么"

❹ 结尾金句，我会不会截图？
   - 不截图 → 重写

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【去AI化深度清洗——逐项检查，必须全部通过】
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

这是口播稿，观众一听就知道是不是AI写的。把所有AI味彻底洗掉。

▌词汇层清查（发现即替换）：
❌ AI高频词：此外、总之、综上所述、值得注意的是、至关重要、深入探讨、深刻的启示、
   引人深省、不禁让人、持久的影响、格局、赋能、洞见、值得深思、由此可见、
   不可磨灭、见证了、标志着、体现了、不言而喻、毋庸置疑
→ 直接删除或换成口语表达

❌ 宣传性空话：充满活力的、丰富的、令人叹为观止的、深刻的、颠覆性的
→ 删掉形容词，用具体事实替代

▌结构层清查：
❌ 三段论公式（首先……其次……最后……）→ 打破，改成自然流动的口语叙述
❌ 否定排比（"这不仅仅是……而是……"）→ 直接说结论
❌ 以"这本书"或"作者"开头的句子（超过2句）→ 换成"你""我"的视角

▌节奏层清查：
❌ 超过30字的长句 → 拆成两句
❌ 连续3句以上相同句型（都是陈述句/都是问句）→ 换节奏
✅ 检查：有没有几个短句打节奏的段落？（好的口播稿一定有）

▌真实感注入：
- 如果全文没有一个具体数字或细节 → 补上（从书中找）
- 如果全文没有"你"字 → 至少加3处，直接和观众说话
- 如果结尾感觉是"总结"而不是"留白/回味" → 改成一句有态度的话

【原稿】
{draft}

【输出要求】
直接输出修改后的完整推荐词。不要说"已修改如下"，不要有任何前缀。
保留原有信息量，让语言更自然、更有力、更像真人在说话。
这篇文章应该让人读完觉得：这是一个真正读过这本书、有自己想法的人写的。"""

    polished, _model = ai_extract(
        system_prompt=PUSHUTREE_SYSTEM_PROMPT,
        user_prompt=user_prompt,
        max_tokens=3000,
        temperature=0.6,
    )
    return polished.strip()


def run_pushutree_task(script_id: str, source_id: str, episode_count: int,
                       platform: str, style: str, direct_text: str = "", direct_book_name: str = ""):
    """后台任务：朴树之道三步流程
    支持两种模式：
    1. 书库模式：source_id 非空，从数据库读取书籍
    2. 独立模式：source_id 为空，用 direct_text + direct_book_name
    """
    conn = get_db()

    def upd(status, progress, message, episodes=None, plan=None, error_msg=None):
        fields = "status=?,progress=?,message=?,updated_at=?"
        vals = [status, progress, message, now()]
        if episodes is not None:
            fields += ",episodes=?"
            vals.append(json.dumps(episodes, ensure_ascii=False))
        if plan is not None:
            fields += ",plan=?"
            vals.append(json.dumps(plan, ensure_ascii=False))
        if error_msg is not None:
            fields += ",error_msg=?"
            vals.append(error_msg)
        vals.append(script_id)
        conn.execute(f"UPDATE scripts SET {fields} WHERE id=?", vals)
        conn.commit()

    try:
        upd("processing", 5, "准备书籍内容...")

        text = ""
        book_name = direct_book_name

        if direct_text:
            # 独立模式：直接用传入的文本
            text = direct_text
            upd("processing", 15, f"已获取书籍文本，共{len(text)}字...")
        elif source_id:
            # 书库模式：从数据库读取
            upd("processing", 8, "读取书籍信息...")
            src = conn.execute("SELECT * FROM sources WHERE id=?", (source_id,)).fetchone()
            if not src:
                upd("error", 0, "找不到书籍", error_msg="source not found")
                return
            book_name = src["title"]

            # 从PDF文件提取文本（任意状态的书籍都可以，不必须提炼完成）
            print(f"[DEBUG] pushutree source: type={src['type']}, file_path={src['file_path']}, url={src['url']}")
            if src["type"] == "book" and src["file_path"]:
                upd("processing", 10, "提取PDF文本...")
                text, _, _ = extract_text_from_pdf(src["file_path"])
                upd("processing", 15, f"PDF文本提取完成（{len(text)}字）...")
            elif src["type"] == "epub" and src["file_path"]:
                # epub 类型
                upd("processing", 10, "解析EPUB文件...")
                text, _, _ = extract_text_from_epub(src["file_path"])
                if isinstance(text, list):
                    text = "\n".join(text)
                upd("processing", 15, f"EPUB解析完成（{len(text)}字）...")
            elif src["type"] == "txt" and src["file_path"]:
                # txt 类型
                upd("processing", 10, "读取文本文件...")
                text, _, _ = extract_text_from_txt(src["file_path"])
                if isinstance(text, list):
                    text = "\n".join(text)
                upd("processing", 15, f"文本文件读取完成（{len(text)}字）...")
            elif src["type"] == "text":
                # text类型：文本存在url字段
                text = src["url"] or ""
                upd("processing", 15, f"已获取文本内容（{len(text)}字）...")
            elif src["type"] == "url":
                # url类型：需要从URL抓取内容
                import asyncio
                loop = asyncio.new_event_loop()
                try:
                    upd("processing", 10, "正在从URL抓取内容...")
                    text = loop.run_until_complete(extract_text_from_url(src["url"]))
                    upd("processing", 15, f"URL内容获取完成（{len(text)}字）...")
                finally:
                    loop.close()
            elif src["url"]:
                # book类型但没有file_path，有url字段——尝试当作URL处理
                import asyncio
                loop = asyncio.new_event_loop()
                try:
                    upd("processing", 10, "正在从URL抓取内容...")
                    text = loop.run_until_complete(extract_text_from_url(src["url"]))
                    upd("processing", 15, f"URL内容获取完成（{len(text)}字）...")
                finally:
                    loop.close()
            else:
                # 兜底：尝试从素材库获取内容
                mat_rows = conn.execute("SELECT content FROM materials WHERE source_id=? LIMIT 1", (source_id,)).fetchall()
                if mat_rows:
                    text = " ".join([m["content"] for m in mat_rows])
                    upd("processing", 15, f"从素材库获取内容（{len(text)}字）...")
                else:
                    upd("error", 0, "无法获取书籍内容", error_msg="no content found")
                    return
        else:
            upd("error", 0, "缺少书籍内容", error_msg="no source")
            return

        if not text or len(text.strip()) < 50:
            print(f"[DEBUG] text too short: len={len(text) if text else 0}, first 100 chars: {text[:100] if text else 'empty'}")
            # 文本太短，尝试从素材库获取内容作为兜底
            if source_id:
                mat_rows = conn.execute("SELECT content FROM materials WHERE source_id=?", (source_id,)).fetchall()
                print(f"[DEBUG] materials from DB: {len(mat_rows)} rows")
                if mat_rows:
                    text = " ".join([m["content"] for m in mat_rows])
                    upd("processing", 15, f"从素材库获取内容（{len(text)}字）...")
            if not text or len(text.strip()) < 50:
                upd("error", 0, "书籍文本太少，无法生成", error_msg="text too short")
                return

        if not book_name:
            book_name = "本书"

        # 第一步：策划
        upd("processing", 20, f"第一步：分析《{book_name}》，生成{episode_count}期策划...")
        plan = pushutree_plan(book_name, text, episode_count, style)
        if not plan:
            upd("error", 0, "策划生成失败，请重试", error_msg="plan empty")
            return
        upd("processing", 35, f"策划完成，共{len(plan)}期，开始撰写...", plan=plan)

        # 第二步+第三步：逐期撰写+精修
        episodes = []
        for i, ep in enumerate(plan):
            ep_no = ep.get("ep_no", i + 1)
            title = ep.get("title", f"第{ep_no}期")

            progress = 35 + int((i / len(plan)) * 55)
            upd("processing", progress, f"撰写第{ep_no}/{len(plan)}期：{title}...")

            try:
                draft = pushutree_write(book_name, text, ep, style)
                upd("processing", progress + 3, f"精修第{ep_no}期...")
                final = pushutree_polish(book_name, draft, ep)
            except Exception as e:
                final = f"（生成失败：{e}）"

            episodes.append({
                "ep_no": ep_no,
                "title": title,
                "content": final,
                "pain_point": ep.get("pain_point", ""),
                "angle": ep.get("angle", ""),
            })
            upd("processing", progress + 5, f"第{ep_no}期完成", episodes=episodes)

        upd("done", 100, f"全部{len(episodes)}期生成完毕！", episodes=episodes)

    except Exception as e:
        import traceback
        upd("error", 0, f"生成失败：{e}", error_msg=traceback.format_exc())
    finally:
        conn.close()


class PushutreeRequest(BaseModel):
    source_id: str = ""           # 可选：已有书籍ID
    book_name: str = ""           # 可选：直接传书名（独立模式）
    book_text: str = ""           # 可选：直接传书籍文本（独立模式）
    episode_count: int = 8
    platform: str = "抖音/视频号"
    style: str = "犀利、接地气、直击痛点"


@app.post("/api/pushutree/create")
def create_pushutree(req: PushutreeRequest):
    """创建朴树之道分享任务
    支持两种模式：
    1. 书库模式：传 source_id（任意状态的书籍都可以，不必须提炼完成）
    2. 独立模式：传 book_name + book_text（直接输入书名和文本）
    """
    conn = get_db()

    # 确定来源信息
    if req.source_id:
        src = conn.execute("SELECT * FROM sources WHERE id=?", (req.source_id,)).fetchone()
        if not src:
            conn.close()
            raise HTTPException(404, "书籍不存在")
        source_id = req.source_id
        source_title = src["title"]
    elif req.book_name:
        # 独立模式：没有对应的 sources 记录，用 book_name 作标题，source_id 为空
        source_id = ""
        source_title = req.book_name
    else:
        conn.close()
        raise HTTPException(400, "请提供书籍ID或书名")

    script_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO scripts VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (script_id, source_id, source_title,
         req.episode_count, req.platform, req.style,
         "pending", 0, "等待生成...",
         "[]", "[]", None, now(), now())
    )
    conn.commit()
    conn.close()

    # 后台异步执行
    threading.Thread(
        target=run_pushutree_task,
        args=(script_id, source_id, req.episode_count, req.platform, req.style, req.book_text, req.book_name),
        daemon=True
    ).start()

    return {"script_id": script_id, "status": "pending"}


@app.get("/api/pushutree/{script_id}")
def get_pushutree(script_id: str):
    """查询朴树之道任务进度和结果"""
    conn = get_db()
    row = conn.execute("SELECT * FROM scripts WHERE id=?", (script_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "任务不存在")
    d = dict(row)
    d["episodes"] = json.loads(d["episodes"] or "[]")
    d["plan"] = json.loads(d["plan"] or "[]")
    return d


@app.get("/api/pushutree")
def list_pushutree(source_id: str = None):
    """获取所有朴树之道系列（可按书籍过滤）"""
    conn = get_db()
    if source_id:
        rows = conn.execute(
            "SELECT * FROM scripts WHERE source_id=? ORDER BY created_at DESC",
            (source_id,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM scripts ORDER BY created_at DESC"
        ).fetchall()
    conn.close()
    result = []
    for row in rows:
        d = dict(row)
        d["episodes"] = json.loads(d["episodes"] or "[]")
        d["plan"] = json.loads(d["plan"] or "[]")
        result.append(d)
    return result


@app.delete("/api/pushutree/{script_id}")
def delete_pushutree(script_id: str):
    """删除朴树之道系列"""
    conn = get_db()
    conn.execute("DELETE FROM scripts WHERE id=?", (script_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.post("/api/pushutree/upload-and-create")
async def pushutree_upload_and_create(
    file: UploadFile = File(None),
    book_name: str = Form(...),
    book_text: str = Form(""),
    episode_count: int = Form(8),
    platform: str = Form("抖音/视频号"),
    style: str = Form("犀利、接地气、直击痛点"),
):
    """朴树之道独立上传入口：
    直接上传PDF或填入书名+文本，不依赖书库，立即生成系列分享文案。
    """
    text = book_text

    # 如果上传了PDF，提取文本
    if file and file.filename:
        file_id = str(uuid.uuid4())
        save_path = UPLOAD_DIR / f"pt_{file_id}_{file.filename}"
        content = await file.read()
        with open(save_path, "wb") as f:
            f.write(content)
        try:
            extracted, _, _ = extract_text_from_pdf(str(save_path))
            text = extracted
        except Exception as e:
            raise HTTPException(500, f"PDF解析失败：{e}")

    if not text or len(text.strip()) < 50:
        raise HTTPException(400, "请上传PDF文件或填入书籍文本（至少50字）")

    if not book_name.strip():
        raise HTTPException(400, "请填写书名")

    # 创建 script 记录
    conn = get_db()
    script_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO scripts VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (script_id, "", book_name.strip(),
         episode_count, platform, style,
         "pending", 0, "等待生成...",
         "[]", "[]", None, now(), now())
    )
    conn.commit()
    conn.close()

    # 后台异步执行
    threading.Thread(
        target=run_pushutree_task,
        args=(script_id, "", episode_count, platform, style, text, book_name.strip()),
        daemon=True
    ).start()

    return {"script_id": script_id, "status": "pending", "book_name": book_name.strip()}


# 统计数据

@app.get("/api/stats")
def get_stats():
    conn = get_db()
    sources_total = conn.execute("SELECT COUNT(*) FROM sources WHERE status='done'").fetchone()[0]
    materials_total = conn.execute("SELECT COUNT(*) FROM materials").fetchone()[0]
    cat_counts = dict(conn.execute(
        "SELECT category, COUNT(*) FROM materials GROUP BY category"
    ).fetchall())
    starred = conn.execute("SELECT COUNT(*) FROM materials WHERE is_starred=1").fetchone()[0]
    conn.close()
    return {
        "sources": sources_total,
        "materials": materials_total,
        "starred": starred,
        "by_category": cat_counts,
    }

# 保存创作
class SaveCreationRequest(BaseModel):
    title: str
    content: str
    platform: str = ""
    source_ids: list = []
    material_ids: list = []

@app.post("/api/creations")
def save_creation(req: SaveCreationRequest):
    cid = str(uuid.uuid4())
    conn = get_db()
    conn.execute(
        "INSERT INTO creations VALUES (?,?,?,?,?,?,?,?)",
        (cid, req.title, req.content,
         json.dumps(req.source_ids), json.dumps(req.material_ids),
         req.platform, now(), now())
    )
    conn.commit()
    conn.close()
    return {"id": cid}

@app.get("/api/creations")
def list_creations(q: str = ""):
    conn = get_db()
    sql = "SELECT * FROM creations"
    params = []
    if q:
        sql += " WHERE title LIKE ? OR content LIKE ?"
        params = [f"%{q}%", f"%{q}%"]
    sql += " ORDER BY updated_at DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]



# process_source_task_smart 已迁移到 services/process.py


# API: 启动智能提取（新版）
@app.post("/api/sources/{source_id}/extract-smart")
def start_smart_extraction(source_id: str, mode: str = "full"):
    """启动智能提取流程（分层Chunking + 多轮Pipeline + 质量评分）"""
    if not _SMART_EXTRACTION_AVAILABLE:
        raise HTTPException(400, "智能提取模块未加载，请检查依赖")

    conn = get_db()
    src = conn.execute("SELECT * FROM sources WHERE id=?", (source_id,)).fetchone()
    if not src:
        conn.close()
        raise HTTPException(404, "来源不存在")

    task_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?)",
        (task_id, source_id, "pending", 0, "等待智能提取...", None, now(), now())
    )
    conn.commit()
    conn.close()

    # 使用新的智能提取流程（传 "smart" 标记，worker 会调用 process_source_task_smart）
    enqueue_task(task_id, source_id, "smart")

    return {
        "task_id": task_id,
        "source_id": source_id,
        "mode": "smart",
        "message": "智能提取任务已创建（分层Chunking+多轮Pipeline+质量评分）"
    }


# API: 获取素材质量详情
@app.get("/api/materials/{material_id}/quality")
def get_material_quality(material_id: str):
    """获取素材的质量评分详情"""
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM materials WHERE id=?", (material_id,)
    ).fetchone()
    conn.close()

    if not row:
        raise HTTPException(404, "素材不存在")

    metadata = json.loads(row["metadata"] or "{}")
    quality_score = metadata.pop("_quality_score", {})

    return {
        "material_id": material_id,
        "quality_score": quality_score,
        "metadata": metadata,
        "suggestions": quality_score.get("suggestions", [])
    }


# API: 对比新旧提取方式（用于评估）
@app.post("/api/sources/{source_id}/extract-compare")
def compare_extraction_methods(source_id: str):
    """对比原版和智能版提取效果（测试用）"""
    # 创建两个任务，分别用两种方式处理
    conn = get_db()
    src = conn.execute("SELECT * FROM sources WHERE id=?", (source_id,)).fetchone()
    if not src:
        conn.close()
        raise HTTPException(404, "来源不存在")

    # 原版任务
    task1_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?)",
        (task1_id, source_id, "pending", 0, "原版提取", None, now(), now())
    )

    # 智能版任务
    task2_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?)",
        (task2_id, source_id, "pending", 0, "智能提取", None, now(), now())
    )
    conn.commit()
    conn.close()

    # 启动任务
    enqueue_task(task1_id, source_id, "full")
    enqueue_task(task2_id, source_id, "full")

    return {
        "classic_task_id": task1_id,
        "smart_task_id": task2_id,
        "message": "对比任务已创建，请分别查看结果"
    }


# ═══════════════════════════════════════════════════════════════════════════
# 跨书整合 API
# ═══════════════════════════════════════════════════════════════════════════

class CrossIntegrateRequest(BaseModel):
    content: str          # 用户输入的素材拼合内容（多本书素材拼在一起）
    mode: str = "主题关联"  # 发现主题关联 / 生成系列策划 / 提炼思维模型

@app.post("/api/studio/cross-integrate")
def cross_integrate(req: CrossIntegrateRequest):
    """
    跨书整合：对来自多本书的素材进行跨书深度分析。
    三种模式：
      - 主题关联：发现多本书共同的深层规律与底层逻辑
      - 系列策划：基于共同主题生成可执行的系列内容策划
      - 思维模型：从多本书中提炼可复用的思维模型框架
    """
    mode_prompts = {
        "主题关联": """你是一位顶级知识整合专家。
用户输入了来自多本不同书籍的素材片段，请完成以下分析：

【任务：发现跨书主题关联】

1. **共同主题识别**（3-5个）
   - 找出这些素材中反复出现的核心主题词或概念
   - 每个主题用一句话点明其在所有书中的共同表达方式

2. **底层规律提炼**
   - 这些书在某个更深的层面上，都在说同一件事是什么？
   - 用"它们都在说：……"的句式表达

3. **视角差异**
   - 不同书从什么不同角度切入这个共同主题？（一句话概括各书视角）

4. **IP创作价值**
   - 基于这个跨书共同主题，可以做什么内容（一句话给出方向）

输出格式：直接按上述4个标题输出，不要废话，不要总结段落。""",

        "系列策划": """你是一位顶级内容策划专家，擅长打造系列化IP内容。
用户输入了来自多本书的素材，请基于这些素材策划一套系列内容。

【任务：生成系列内容策划】

1. **系列主题**
   - 系列名称（5-10字，有记忆点）
   - 一句话定位（这个系列为谁解决什么问题）

2. **系列结构**（5-8期）
   - 每期标题（能独立传播，又形成系列感）
   - 每期核心观点（一句话）
   - 对应使用哪本书的哪个核心素材

3. **推送节奏建议**
   - 建议发布频率和顺序逻辑（为什么这样排）

4. **爆款切入点**
   - 哪一期最适合作为首发（理由）

直接输出策划内容，不要有任何前言和废话。""",

        "思维模型": """你是一位知识萃取专家，擅长从书籍素材中提炼可复用的思维框架。
用户输入了多本书的核心素材，请从中提炼思维模型。

【任务：提炼跨书思维模型】

1. **核心模型命名**
   - 给这个思维模型起一个好记的名字（可以是"XX法则""XX模型""XX框架"等）
   - 一句话说明这个模型解决什么问题

2. **模型结构**
   - 用2-4个步骤或维度描述这个模型的运作逻辑
   - 每个步骤/维度：名称 + 一句话解释 + 来自哪本书的什么概念

3. **适用场景**
   - 这个思维模型在哪3个场景下最有用？（具体场景，不要抽象）

4. **记忆口诀**（可选）
   - 如果能用一句顺口的话记住这个模型，是什么？

5. **内容创作切入**
   - 用这个思维模型，可以写什么选题？（给出3个具体标题）

直接输出，不要废话，要让人看完立刻能用。"""
    }

    system_prompt = f"""你是一位顶级IP内容创作专家，专注于职场认知升级、人性洞察和个人成长破局三大方向。

{mode_prompts.get(req.mode, mode_prompts['主题关联'])}

【去AI化规则——必须执行】
- 禁止"综上所述""值得注意""深刻启示""引人深省""不言而喻"等AI高频词
- 用真实直接的语言，有判断有态度，不说官话
- 说具体的，不说抽象的"""

    user_prompt = f"以下是来自多本书的素材内容，请按要求分析：\n\n{req.content}"

    try:
        result, model_used = ai_extract(system_prompt, user_prompt, max_tokens=3000, temperature=0.75)
        return {"result": result, "model": model_used, "mode": req.mode}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"跨书整合失败：{str(e)}")


# ═══════════════════════════════════════════════════════════════════════════
# MediaCrawler 数据接入
# ═══════════════════════════════════════════════════════════════════════════

# MediaCrawler 数据目录（本地 media_crawler/data 或从 config.json 读取）
_MEDIA_CRAWLER_LOCAL_DATA = BASE_DIR / "media_crawler" / "data"
_CONFIG_PATH = BASE_DIR / "config.json"
if _CONFIG_PATH.exists():
    try:
        with open(_CONFIG_PATH, "r", encoding="utf-8") as f:
            _cfg = json.load(f)
            _ext_path = _cfg.get("media_crawler", {}).get("data_dir", "")
            MEDIA_CRAWLER_DATA_DIR = Path(_ext_path) if _ext_path else _MEDIA_CRAWLER_LOCAL_DATA
    except Exception:
        MEDIA_CRAWLER_DATA_DIR = _MEDIA_CRAWLER_LOCAL_DATA
else:
    MEDIA_CRAWLER_DATA_DIR = _MEDIA_CRAWLER_LOCAL_DATA

# 平台 ID → 中文名映射
PLATFORM_NAMES = {
    "douyin": "抖音",
    "xhs": "小红书",
    "weibo": "微博",
    "kuaishou": "快手",
    "bili": "B站",
    "tieba": "贴吧",
    "zhihu": "知乎",
}


class MediaImportRequest(BaseModel):
    records: List[dict]  # 每条记录包含 platform, content, metadata


@app.get("/api/media/platforms")
def list_media_platforms():
    """返回 MediaCrawler 已采集的平台列表"""
    if not MEDIA_CRAWLER_DATA_DIR.exists():
        return {"platforms": []}

    platforms = []
    for subdir in MEDIA_CRAWLER_DATA_DIR.iterdir():
        if subdir.is_dir() and subdir.name != "wechat-output":
            pid = subdir.name
            platforms.append({
                "id": pid,
                "name": PLATFORM_NAMES.get(pid, pid),
                "path": str(subdir),
            })

    return {"platforms": platforms}


@app.get("/api/media/files")
def list_media_files(platform: str = ""):
    """返回指定平台的数据文件列表"""
    if not platform:
        return {"files": []}

    platform_dir = MEDIA_CRAWLER_DATA_DIR / platform
    if not platform_dir.exists():
        return {"files": []}

    files = []
    for subdir in platform_dir.iterdir():
        if subdir.is_dir():
            for f in subdir.iterdir():
                if f.suffix in (".json", ".jsonl", ".csv"):
                    files.append({
                        "name": f.name,
                        "path": str(f.relative_to(MEDIA_CRAWLER_DATA_DIR)).replace("\\", "/"),
                        "size": f.stat().st_size,
                    })

    # 也检查顶层文件
    for f in platform_dir.iterdir():
        if f.is_file() and f.suffix in (".json", ".jsonl", ".csv"):
            files.append({
                "name": f.name,
                "path": str(f.relative_to(MEDIA_CRAWLER_DATA_DIR)).replace("\\", "/"),
                "size": f.stat().st_size,
            })

    return {"files": files}


@app.get("/api/media/data")
def get_media_data(platform: str = "", file: str = "", offset: int = 0, limit: int = 20):
    """读取指定文件的数据，支持分页"""
    if not platform or not file:
        return {"records": [], "total": 0, "offset": offset, "limit": limit}

    file_path = MEDIA_CRAWLER_DATA_DIR / file
    if not file_path.exists():
        return {"records": [], "total": 0, "offset": offset, "limit": limit}

    records = []
    try:
        if file_path.suffix == ".jsonl":
            with open(file_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            total = len(lines)
            for line in lines[offset:offset + limit]:
                try:
                    records.append(json.loads(line))
                except:
                    pass
        elif file_path.suffix == ".json":
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                total = len(data)
                records = data[offset:offset + limit]
            else:
                total = 1
                records = [data] if offset == 0 else []
    except Exception as e:
        return {"records": [], "total": 0, "offset": offset, "limit": limit, "error": str(e)}

    return {"records": records, "total": total, "offset": offset, "limit": limit}


@app.post("/api/media/import")
def import_media_records(req: MediaImportRequest):
    """将选中的 MediaCrawler 记录导入到 ip-arsenal 素材库"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    imported = 0
    failed = 0
    today = datetime.now().strftime("%Y-%m-%d")

    for record in req.records:
        try:
            platform = record.get("platform", "unknown")
            content = record.get("content", "") or record.get("desc", "") or ""
            if not content:
                failed += 1
                continue

            metadata = record.get("metadata", {})
            if not isinstance(metadata, dict):
                metadata = {}

            # 从 record 直接提取有用字段放入 metadata
            for key in ("nickname", "liked_count", "comment_count", "share_count",
                        "collect_count", "user_id", "avatar", "ip_location",
                        "source_keyword", "create_time", "note_id", "aweme_id",
                        "note_url", "aweme_url"):
                if key in record and key not in metadata:
                    metadata[key] = record[key]

            source_id = f"media_crawler_{platform}_{today}"

            # 插入 materials 表
            mid = str(uuid.uuid4())
            c.execute("""
                INSERT INTO materials (id, source_id, category, content, metadata, tags, platform, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                mid,
                source_id,
                "topic",  # 默认分类，可后续 AI 重新分类
                content,
                json.dumps(metadata, ensure_ascii=False),
                "[]",
                platform,
                datetime.now().isoformat(),
            ))
            imported += 1
        except Exception as e:
            failed += 1
            print(f"[MediaImport] 导入失败: {e}")

    conn.commit()
    conn.close()

    return {"imported": imported, "failed": failed}


# ═══════════════════════════════════════════════════════════════════════════
# MediaCrawler API 代理（方案B：深度融合）
# ip-arsenal 后端作为网关，转发请求到 MediaCrawler (port 8080)
# ═══════════════════════════════════════════════════════════════════════════

MEDIACRAWLER_BASE = "http://localhost:8080"


@app.get("/api/mediacrawler/health")
async def mc_health():
    """转发到 MediaCrawler /api/health"""
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.get(f"{MEDIACRAWLER_BASE}/api/health")
        return r.json()


@app.get("/api/mediacrawler/config/platforms")
async def mc_config_platforms():
    """转发到 MediaCrawler /api/config/platforms"""
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.get(f"{MEDIACRAWLER_BASE}/api/config/platforms")
        return r.json()


@app.get("/api/mediacrawler/config/options")
async def mc_config_options():
    """转发到 MediaCrawler /api/config/options"""
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.get(f"{MEDIACRAWLER_BASE}/api/config/options")
        return r.json()


@app.post("/api/mediacrawler/crawler/start")
async def mc_crawler_start(req: dict):
    """转发到 MediaCrawler /api/crawler/start"""
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.post(f"{MEDIACRAWLER_BASE}/api/crawler/start", json=req)
        return r.json()


@app.post("/api/mediacrawler/crawler/stop")
async def mc_crawler_stop():
    """转发到 MediaCrawler /api/crawler/stop"""
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.post(f"{MEDIACRAWLER_BASE}/api/crawler/stop")
        return r.json()


@app.get("/api/mediacrawler/crawler/status")
async def mc_crawler_status():
    """转发到 MediaCrawler /api/crawler/status"""
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.get(f"{MEDIACRAWLER_BASE}/api/crawler/status")
        return r.json()


@app.get("/api/mediacrawler/crawler/logs")
async def mc_crawler_logs(limit: int = 100):
    """转发到 MediaCrawler /api/crawler/logs"""
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.get(f"{MEDIACRAWLER_BASE}/api/crawler/logs", params={"limit": limit})
        return r.json()


@app.get("/api/mediacrawler/data/files")
def mc_data_files(platform: str = ""):
    """返回本地 MediaCrawler 数据目录的文件（不走代理，避免跨域问题）"""
    return list_media_files(platform=platform)


@app.get("/api/mediacrawler/data/files/{file_path:path}")
def mc_data_file_preview(file_path: str, preview: bool = True, limit: int = 50):
    """从本地 MediaCrawler 数据目录读取文件内容"""
    # file_path 格式：douyin/jsonl/search_contents_2026-03-30.jsonl
    fp = MEDIA_CRAWLER_DATA_DIR / file_path
    if not fp.exists():
        return {"records": [], "total": 0, "offset": 0, "limit": limit, "error": "file not found: " + str(fp)}

    if preview:
        # 直接读取文件内容，避免调用 get_media_data 的 platform/file 拼接逻辑
        records = []
        total = 0
        try:
            if fp.suffix == ".jsonl":
                with open(fp, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                total = len(lines)
                for line in lines[0:limit]:
                    try:
                        records.append(json.loads(line))
                    except:
                        pass
            elif fp.suffix == ".json":
                with open(fp, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    total = len(data)
                    records = data[0:limit]
                else:
                    total = 1
                    records = [data]
        except Exception as e:
            return {"records": [], "total": 0, "offset": 0, "limit": limit, "error": str(e)}
        return {"records": records, "total": total, "offset": 0, "limit": limit}
    else:
        with open(fp, "r", encoding="utf-8") as f:
            content = f.read()
        return {"content": content}


@app.websocket("/api/mediacrawler/ws/status")
async def mc_ws_status(websocket: WebSocket):
    """WebSocket 转发：/api/ws/status"""
    await websocket.accept()
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            async with client.stream("GET", f"{MEDIACRAWLER_BASE}/api/ws/status") as resp:
                async for text in resp.aiter_text():
                    if text:
                        await websocket.send_text(text)
    except Exception as e:
        print(f"[mc_ws_status] error: {e}")
    finally:
        await websocket.close()


# 静态文件（前端）—— 只在非 API 路径上提供服务（放最后，因为要注册在 API 路由之后）
@app.get("/{path:path}")
async def serve_frontend(path: str):
    from fastapi import HTTPException
    from fastapi.responses import FileResponse
    if path.startswith("api/"):
        raise HTTPException(status_code=404, detail="Not Found")
    file_path = FRONT_DIR / path
    if file_path.is_file():
        return FileResponse(file_path)
    # 其它路径返回 index.html（SPA 路由）
    return FileResponse(FRONT_DIR / "index.html")


if __name__ == "__main__":
    import uvicorn, sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    print("IP Arsenal starting on http://localhost:8766")
    print(f"[SmartExtraction] 智能提取模块: {'可用' if _SMART_EXTRACTION_AVAILABLE else '不可用'}")
    # Worker 线程和任务恢复由 @app.on_event("startup") 统一处理
    uvicorn.run("main:app", host="0.0.0.0", port=8766, reload=False)

